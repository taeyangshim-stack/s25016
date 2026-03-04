"""
회의록 + 아젠다 Excel 생성 스크립트 (A4 인쇄 최적화)
"""
import openpyxl
from pathlib import Path
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

_HERE = Path(__file__).parent

# ──────────────────────────────────────────────
# 공통 헬퍼
# ──────────────────────────────────────────────
def thin():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def medium():
    s = Side(style='medium', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def medium_box(top=True, bottom=True, left=True, right=True):
    m = Side(style='medium', color='000000')
    t = Side(style='thin', color='AAAAAA')
    return Border(
        left=m if left else t,
        right=m if right else t,
        top=m if top else t,
        bottom=m if bottom else t,
    )

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell(ws, row, col, value, bold=False, size=10, color='000000',
         bg=None, align='left', valign='center', wrap=True, border=None, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    if border is not None:
        c.border = border
    return c

def merge(ws, r1, c1, r2, c2, value, bold=False, size=10, color='000000',
          bg=None, align='left', valign='center', wrap=True, border=None, italic=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    if border is not None:
        c.border = border
    return c

def set_border_range(ws, r1, c1, r2, c2, border):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = border

def a4_setup(ws, landscape=False):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.7, bottom=0.7,
                                  header=0.3, footer=0.3)
    ws.print_options.gridLines = False

# ══════════════════════════════════════════════════════════════════
#  파일 1: 회의록 (KUKA·우리웰텍, 2026-03-03)
# ══════════════════════════════════════════════════════════════════
def make_minutes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "회의록"
    a4_setup(ws, landscape=False)

    # 열 너비 (A~G, 총 7열)
    col_w = [2, 22, 16, 20, 20, 16, 4]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ROW = 1

    # ── 문서 제목 ──────────────────────────────
    ws.row_dimensions[ROW].height = 36
    merge(ws, ROW, 2, ROW, 6,
          'ABB 현안 검토 및 KUKA 기술 적용 가능성 회의록',
          bold=True, size=16, color='FFFFFF', bg='1A3A5C',
          align='center', valign='center', border=medium())
    ROW += 1

    ws.row_dimensions[ROW].height = 16
    merge(ws, ROW, 2, ROW, 6,
          '삼성중공업 34bay 자동용접 갠트리 로봇 시스템 — 후속 라인 로봇 메이커 검토',
          size=9, color='555555', bg='EAF2FB', align='center', border=thin())
    ROW += 1
    ROW += 1  # 공백

    # ── 섹션 헬퍼 ─────────────────────────────
    def section(title, r):
        ws.row_dimensions[r].height = 22
        merge(ws, r, 2, r, 6, title, bold=True, size=11,
              color='FFFFFF', bg='1A5276', align='left', border=medium())
        return r + 1

    def row2(ws, r, label, value):
        ws.row_dimensions[r].height = 18
        cell(ws, r, 2, label, bold=True, size=9, bg='D6EAF8',
             align='center', border=thin())
        merge(ws, r, 3, r, 6, value, size=9, border=thin())
        return r + 1

    # ── 1. 회의 개요 ──────────────────────────
    ROW = section('1. 회의 개요', ROW)
    ROW = row2(ws, ROW, '일시', '2026년 3월 3일 (화)  13:00 ~ 14:30  (1시간 30분)')
    ROW = row2(ws, ROW, '장소', 'SP시스템스 어곡본사 회의실')
    ROW = row2(ws, ROW, '안건', '① ABB 현안 및 KUKA 기술 상황 검토     ② 인도 협력 투자 관련 논의')
    ROW = row2(ws, ROW, '작성자', '심태양 부장 (SP시스템스)')
    ROW += 1

    # ── 2. 참석자 ─────────────────────────────
    ROW = section('2. 참석자', ROW)
    ws.row_dimensions[ROW].height = 18
    for col, hdr in enumerate(['소속', '성명', '직책', '연락처'], 2):
        cell(ws, ROW, col, hdr, bold=True, size=9, bg='D6EAF8',
             align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=5, end_row=ROW, end_column=6)
    cell(ws, ROW, 5, '연락처', bold=True, size=9, bg='D6EAF8',
         align='center', border=thin())
    ROW += 1

    attendees = [
        ('SP시스템스', '심태양', '부장', '—'),
        ('우리웰텍', '이희덕', '대표이사', '010-3594-5085'),
        ('우리웰텍', '한창헌', '기술고문', '010-5874-0289'),
        ('KUKA 로보틱스코리아', '김형석', '부장 / 기술지원팀', '010-4453-1451'),
    ]
    for so, nm, pos, tel in attendees:
        ws.row_dimensions[ROW].height = 17
        cell(ws, ROW, 2, so, size=9, align='center', border=thin())
        cell(ws, ROW, 3, nm, size=9, bold=True, align='center', border=thin())
        cell(ws, ROW, 4, pos, size=9, align='center', border=thin())
        ws.merge_cells(start_row=ROW, start_column=5, end_row=ROW, end_column=6)
        cell(ws, ROW, 5, tel, size=9, align='center', border=thin())
        ROW += 1
    ROW += 1

    # ── 3. 안건 1 ─────────────────────────────
    ROW = section('3. 안건 1 — ABB 현안 및 KUKA 기술 검토', ROW)

    def speaker_block(ws, r, who, badge_bg, items, note=None):
        ws.row_dimensions[r].height = 17
        merge(ws, r, 2, r, 6, f'▶ {who}', bold=True, size=9,
              color='FFFFFF', bg=badge_bg, border=thin())
        r += 1
        for itm in items:
            ws.row_dimensions[r].height = 30
            cell(ws, r, 2, '•', size=9, bg='FDFEFE', align='center', border=thin())
            merge(ws, r, 3, r, 6, itm, size=9, bg='FDFEFE', border=thin(), wrap=True)
            r += 1
        if note:
            ws.row_dimensions[r].height = 30
            merge(ws, r, 2, r, 6, f'⚠ {note}', size=9, bg='FDF2F2',
                  color='922B21', border=thin(), wrap=True)
            r += 1
        return r

    ROW = speaker_block(ws, ROW, 'SP시스템스  심태양 부장', '2980B9', [
        '삼성중공업 거제조선소 34bay 자동용접 갠트리 로봇 시스템 컨셉 및 로봇 Config 방법 설명',
        'ABB Korea 지원 하에 삼성중공업·SP시스템스가 자체적으로 현안 해결 진행 중',
        '핵심 이슈: ABB IRC5 컨트롤러 단종 → 신규 컨트롤러와 Lincoln 용접기 인터페이스 기술 성능 확인 필요',
        '기존 라인의 문제가 신규 라인에서도 동일하게 대두될 가능성 있음',
    ], note='ABB IRC5 단종 이슈 — 신규 컨트롤러 도입 시 Lincoln 용접기 인터페이스 및 성능 검증 필요. 3/4(수) ABB·삼성·SP 3자 미팅 확인 예정.')

    ROW = speaker_block(ws, ROW, 'KUKA 로보틱스코리아  김형석 부장', 'D35400', [
        'SyncMove 기능 적용으로 갠트리 동기 제어 대응 가능할 것으로 판단',
        '갠트리 X·Y·Z·R 4축을 단일 갠트리로 Config하여 시스템 구성 가능하다는 의견',
        '신규 라인 적용 여부에 대해 이번 주 금요일(3/7)까지 회신 예정 → 전향적 검토 또는 계약 조건 기반 검토 방향',
    ])

    ROW = speaker_block(ws, ROW, '우리웰텍  이희덕 대표 · 한창헌 고문', '27AE60', [
        '메이저 로봇 업체는 대부분의 어플리케이션 대응 가능',
        '엔지니어링 가능한 메이커와 함께 일하는 것이 중요 — 본사 중심 대응 구조는 문제 해결 효율 저하',
        'SP시스템스의 검토 과정에 이 관점 반영 요청',
    ])

    ROW = speaker_block(ws, ROW, 'SP시스템스  심태양 부장 — 요청 사항', '1F618D', [
        '신규 라인에 대해 KUKA의 전향적·긍정적 검토 강력 요청',
        '타 메이커 검토를 위해서는 정량적 가능성 판단 자료가 반드시 필요 → KUKA 측에서 도출·제공 요청',
        '신규 라인에 대한 전향적 검토를 재차 강조하며 적극적 협력 요청',
    ])
    ROW += 1

    # ── 4. 안건 2 ─────────────────────────────
    ROW = section('4. 안건 2 — 인도 협력 투자 관련', ROW)
    ws.row_dimensions[ROW].height = 30
    merge(ws, ROW, 2, ROW, 6,
          '우리웰텍 인도 시장 협력 관계를 통해 트라이 케이스 다수 보유.',
          size=9, bg='FDFEFE', border=thin(), wrap=True)
    ROW += 1
    ws.row_dimensions[ROW].height = 30
    merge(ws, ROW, 2, ROW, 6,
          '별도 논의: SP시스템스 영업 송은민 부장 ↔ 우리웰텍 이희덕 대표 간 인도 협력 방향 논의 진행. 구체적 협력 방안은 추가 미팅에서 확정 예정.',
          size=9, bg='FDF9E7', border=thin(), wrap=True)
    ROW += 1
    ROW += 1

    # ── 5. 팔로업 ─────────────────────────────
    ROW = section('5. 팔로업 액션', ROW)
    ws.row_dimensions[ROW].height = 18
    for col, hdr, w in [
        (2, '기한', None), (3, '담당', None), (4, '내용', None),
        (5, '내용(계속)', None), (6, '상태', None)
    ]:
        cell(ws, ROW, col, hdr, bold=True, size=9, bg='2C3E50',
             color='FFFFFF', align='center', border=thin())
    # 내용 열 합치기 헤더
    ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
    ROW += 1

    actions = [
        ('3/4 (수)', 'SP시스템스 · ABB · 삼성중공업',
         'IRC5 후속 컨트롤러 + Lincoln 용접기 인터페이스 기술 성능 확인 3자 미팅', '예정'),
        ('3/7 (금)', 'KUKA 김형석 부장',
         '신규 라인 전향적 검토 또는 계약 기반 검토 회신 (정량적 자료 포함)', '대기'),
        ('별도 협의', '송은민 부장 · 이희덕 대표',
         '인도 협력 투자 방향 후속 논의', '예정'),
    ]
    for dt, who, what, status in actions:
        ws.row_dimensions[ROW].height = 30
        cell(ws, ROW, 2, dt, size=9, align='center', border=thin())
        cell(ws, ROW, 3, who, size=9, align='center', border=thin())
        ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
        cell(ws, ROW, 4, what, size=9, border=thin(), wrap=True)
        cell(ws, ROW, 6, status, size=9, align='center',
             bg='FADBD8', color='922B21', bold=True, border=thin())
        ROW += 1
    ROW += 1

    # ── 6. 요약 ───────────────────────────────
    ROW = section('6. 요약 및 시사점', ROW)
    summary = [
        'ABB IRC5 단종 이슈가 신규 라인 로봇 메이커 검토의 트리거로 작용',
        'KUKA는 기술적으로 SyncMove + 갠트리 Config로 대응 가능 판단 — 3/7(금) 정식 회신 예정',
        '우리웰텍은 로컬 엔지니어링 역량을 메이커 선정의 핵심 기준으로 제시',
        'SP시스템스는 KUKA 측에 정량적 검토 자료 제출 요청 → 의사결정 근거 확보 필요',
        '내일(3/4) ABB 미팅 결과가 신규 라인 방향 결정의 중요한 분기점이 될 것으로 예상',
    ]
    for itm in summary:
        ws.row_dimensions[ROW].height = 20
        cell(ws, ROW, 2, '■', size=9, bg='EAF4FD', align='center', border=thin())
        merge(ws, ROW, 3, ROW, 6, itm, size=9, bg='EAF4FD', border=thin())
        ROW += 1

    ROW += 1
    ws.row_dimensions[ROW].height = 16
    merge(ws, ROW, 2, ROW, 6,
          '다음 일정: [3/4 수] ABB·삼성중공업·SP 기술 미팅   /   [3/7 금] KUKA 김형석 부장 회신 수령',
          size=9, bold=True, color='1A5276', bg='D6EAF8',
          align='center', border=thin())
    ROW += 1
    ROW += 1

    # 푸터
    merge(ws, ROW, 2, ROW, 6,
          '작성: 심태양 부장 (SP시스템스)   |   작성일: 2026-03-03   |   문서번호: S25016-MOM-260303',
          size=8, color='888888', italic=True, align='center', border=thin())

    wb.save(_HERE / '260303_KUKA_우리웰텍_ABB현안_회의록.xlsx')
    print('회의록 저장 완료')


# ══════════════════════════════════════════════════════════════════
#  파일 2: 미팅 아젠다 (ABB·삼성·SP, 2026-03-04)
# ══════════════════════════════════════════════════════════════════
def make_agenda():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "미팅아젠다"
    a4_setup(ws, landscape=False)

    col_w = [2, 14, 28, 16, 16, 14, 4]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ROW = 1

    # 제목
    ws.row_dimensions[ROW].height = 36
    merge(ws, ROW, 2, ROW, 6,
          'ABB · 삼성중공업 · SP시스템스  기술 협의 미팅 아젠다',
          bold=True, size=16, color='FFFFFF', bg='1B2631',
          align='center', valign='center', border=medium())
    ROW += 1
    ws.row_dimensions[ROW].height = 15
    merge(ws, ROW, 2, ROW, 6,
          '삼성중공업 34bay B라인 자동용접 로봇 — 유지보수 현안 및 향후 대응 방안',
          size=9, color='555555', bg='D5D8DC', align='center', border=thin())
    ROW += 1
    ROW += 1

    # 개요
    def section(title, r):
        ws.row_dimensions[r].height = 22
        merge(ws, r, 2, r, 6, title, bold=True, size=11,
              color='FFFFFF', bg='1A3A5C', align='left', border=medium())
        return r + 1

    def row2(ws, r, label, value):
        ws.row_dimensions[r].height = 18
        cell(ws, r, 2, label, bold=True, size=9, bg='D6EAF8',
             align='center', border=thin())
        merge(ws, r, 3, r, 6, value, size=9, border=thin())
        return r + 1

    ROW = section('1. 미팅 개요', ROW)
    ROW = row2(ws, ROW, '일시', '2026년 3월 4일 (수)  14:00')
    ROW = row2(ws, ROW, '장소', '삼성중공업 거제조선소 자동화 실험동')
    ROW = row2(ws, ROW, '목적', 'ADU 긴급 교체 후 원인 분석, 유지보수 체계 수립, 교육 계획 수립, 잔존 이슈 논의')
    ROW += 1

    # 참석자
    ROW = section('2. 참석자', ROW)
    ws.row_dimensions[ROW].height = 18
    for col, hdr in [(2, '소속'), (3, '성명'), (4, '직책'), (5, '비고')]:
        cell(ws, ROW, col, hdr, bold=True, size=9, bg='D6EAF8',
             align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=5, end_row=ROW, end_column=6)
    ROW += 1

    att = [
        ('SP시스템스', '심태양', '부장 / 개발팀', 'DBEAFE', '1E3A5F'),
        ('SP시스템스', '이상주', '대리 / 제어팀', 'DBEAFE', '1E3A5F'),
        ('SP시스템스', '이마룬', '사원 / 개발팀', 'DBEAFE', '1E3A5F'),
        ('삼성중공업', '이무림', '담당', 'D5F5E3', '1D6A39'),
        ('삼성중공업', '이동혁', '담당', 'D5F5E3', '1D6A39'),
        ('ABB Korea', '김태경', '차장 / Technical Sales', 'FADBD8', '7B241C'),
        ('ABB Korea', '이영재', '부장 / 서비스팀', 'FADBD8', '7B241C'),
        ('ABB Korea', '윤건태', '과장 / 서비스팀', 'FADBD8', '7B241C'),
    ]
    for so, nm, pos, bg_c, fg_c in att:
        ws.row_dimensions[ROW].height = 17
        cell(ws, ROW, 2, so, size=9, bg=bg_c, align='center', border=thin())
        cell(ws, ROW, 3, nm, size=9, bold=True, bg=bg_c, align='center', border=thin())
        cell(ws, ROW, 4, pos, size=9, bg=bg_c, align='center', border=thin())
        ws.merge_cells(start_row=ROW, start_column=5, end_row=ROW, end_column=6)
        cell(ws, ROW, 5, '', size=9, bg=bg_c, border=thin())
        ROW += 1
    ROW += 1

    # 아젠다 항목 헬퍼
    def agenda_hdr(ws, r, num, title, time_str):
        ws.row_dimensions[r].height = 22
        merge(ws, r, 2, r, 5,
              f'아젠다 {num}  |  {title}', bold=True, size=10,
              color='FFFFFF', bg='1F618D', border=medium())
        cell(ws, r, 6, time_str, bold=True, size=9,
             color='FFFFFF', bg='1F618D', align='center', border=medium())
        return r + 1

    def agenda_point(ws, r, text, bg='FDFEFE'):
        ws.row_dimensions[r].height = 30
        cell(ws, r, 2, '•', size=9, bg=bg, align='center', border=thin())
        merge(ws, r, 3, r, 6, text, size=9, bg=bg, border=thin(), wrap=True)
        return r + 1

    def ask_box(ws, r, text):
        ws.row_dimensions[r].height = 35
        cell(ws, r, 2, 'Q', bold=True, size=9, bg='D6EAF8',
             color='1A5276', align='center', border=thin())
        merge(ws, r, 3, r, 6, text, size=9, bg='EAF4FD',
              color='1A5276', border=thin(), wrap=True)
        return r + 1

    def warn_box(ws, r, text):
        ws.row_dimensions[r].height = 35
        cell(ws, r, 2, '⚠', size=9, bg='FADBD8',
             align='center', border=thin())
        merge(ws, r, 3, r, 6, text, size=9, bg='FDF2F2',
              color='922B21', border=thin(), wrap=True)
        return r + 1

    # ── 아젠다 섹션 ──
    ROW = section('3. 아젠다', ROW)

    # 아젠다 1
    ROW = agenda_hdr(ws, ROW, 1, '기존 로봇 유지보수 현황 공유', '10분')
    ROW = agenda_point(ws, ROW, '삼성중공업 34bay B라인 운영 현황 및 최근 이슈 이력 공유')
    ROW = agenda_point(ws, ROW, '현재 유지보수 체계 및 역할 분담 현황 확인')
    ROW = agenda_point(ws, ROW, 'ABB 측 현재 기술 지원 범위 및 대응 체계 재확인')
    ROW += 1

    # 아젠다 2
    ROW = agenda_hdr(ws, ROW, 2, '유지보수 교육 전 긴급 대응 체계 수립', '15분')
    ROW = agenda_point(ws, ROW, '교육 완료 전까지 발생하는 이슈에 대한 ABB 긴급 대응 절차 및 범위 명확화')
    ROW = agenda_point(ws, ROW, '긴급 연락 체계 (유선/방문) 및 응답 시간 기준 협의')
    ROW = agenda_point(ws, ROW, 'ABB Service Agreement 진행 협의 여부 검토')
    ROW = ask_box(ws, ROW, '확인 요청: 교육 완료 전 동일 이슈 재발 시 ABB의 구체적 출동 기준과 SLA(응답 시간)는?')
    ROW += 1

    # 아젠다 3
    ROW = agenda_hdr(ws, ROW, 3, 'ADU 긴급 교체 건 — 원인 분석 및 스페어 확보', '20분')
    ROW = warn_box(ws, ROW,
        '2026-02-25 ADU 고장 → 삼성중공업 자체 스페어로 라인 복구. ABB 측 원인 분석 중.')
    ROW = agenda_point(ws, ROW, 'ADU 고장 원인 분석 결과 공유 (ABB 그룹 진단 결과)')
    ROW = agenda_point(ws, ROW,
        '핵심 질문: 삼성중공업 스페어가 없었다면 ABB는 어떻게 대응할 예정이었는가?',
        bg='FDF9E7')
    ROW = agenda_point(ws, ROW,
        '신규 ADU 보충 일정 확인 — ABB → 삼성중공업 스페어 재제출 (기약속: 3/6 수령 예정)')
    ROW = ask_box(ws, ROW,
        '확인: ① ADU 수령 일정 재확인(3/6 가능 여부)   ② 권고 스페어 목록 제출 요청   ③ ABB 긴급 조달 Lead Time')
    ROW += 1

    # 아젠다 4
    ROW = agenda_hdr(ws, ROW, 4, '유지보수 교육 계획 수립', '20분')
    ROW = agenda_point(ws, ROW,
        '삼성중공업 유지보수 교육 — SP시스템스 부산 녹산 RIMS 장비 활용 가능 (기확인)')
    ROW = agenda_point(ws, ROW,
        '추가 요청: 교육 시 SP시스템스 인원도 함께 교육받을 수 있도록 포함 요청\n'
        '→ 삼성중공업 Key Maintenance Person 2명 + SP 인원 포함 구성',
        bg='FDF9E7')
    ROW = agenda_point(ws, ROW, '교육 장소 우선순위: SP 녹산 RIMS → 불가 시 ABB 천안 공장')
    ROW = agenda_point(ws, ROW, '교육 내용 3사 공동 협의 후 진행')

    # 일정표
    ws.row_dimensions[ROW].height = 15
    merge(ws, ROW, 2, ROW, 6, '교육 가능 시점 — 녹산 이설 완료 이후',
          bold=True, size=9, bg='D5F5E3', color='1D6A39',
          align='center', border=thin())
    ROW += 1
    for col, hdr in [(2, '구분'), (3, '일정'), (4, '내용'), (5, '비고')]:
        ws.row_dimensions[ROW].height = 16
        cell(ws, ROW, col, hdr, bold=True, size=9, bg='A9DFBF',
             align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=5, end_row=ROW, end_column=6)
    ROW += 1
    sched = [
        ('1차 측정', '3/18(수) ~ 3/25(수)', 'X축 레일 정도 측정 및 조정 — 부산 녹산'),
        ('2차 측정', '4/6(월) ~ 4/8(수)', '갠트리 로봇 운전 정도 측정 · SKEW · 공간계측'),
    ]
    for s1, s2, s3 in sched:
        ws.row_dimensions[ROW].height = 18
        cell(ws, ROW, 2, s1, size=9, bg='EAFAF1', align='center', border=thin())
        cell(ws, ROW, 3, s2, size=9, bg='EAFAF1', align='center', border=thin())
        ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=6)
        cell(ws, ROW, 4, s3, size=9, bg='EAFAF1', border=thin())
        ROW += 1
    ws.row_dimensions[ROW].height = 18
    merge(ws, ROW, 2, ROW, 6,
          '→ 교육 가능 시점: 2차 측정 완료 이후 (4월 8일 이후) | 현장 상황에 따라 일정 변동 가능',
          size=9, bold=True, color='1D6A39', bg='D5F5E3',
          align='center', border=thin())
    ROW += 1
    ROW = ask_box(ws, ROW,
        '확인: ① ABB 교육 커리큘럼(안) 사전 제출 가능 여부   ② 4월 중순 이후 일정 가협의 가능 여부')
    ROW += 1

    # 아젠다 5
    ROW = agenda_hdr(ws, ROW, 5, 'ABB 잔존 이슈 — 로봇 기본 기능 관련 (Plan B 별도)', '15분')
    ws.row_dimensions[ROW].height = 20
    merge(ws, ROW, 2, ROW, 6,
          '본 안건은 Plan B 프로그램 개발과 별개로, ABB 로봇 기본 기능 수준에서의 이슈 및 기술 문의입니다.',
          size=9, bg='FDF9E7', color='784212', border=thin(), wrap=True)
    ROW += 1
    ROW = agenda_point(ws, ROW, 'BullsEye (불스아이) — 툴 교정 기능 관련 이슈 및 운용 방법 문의')
    ROW = agenda_point(ws, ROW, 'ArcWare (아크웨어) — 용접 소프트웨어 기능 및 파라미터 관련 이슈')
    ROW = agenda_point(ws, ROW, 'RAPID 문법 관련 — 기본 언어 사용 시 발생하는 문법 오류 및 제약 사항')
    ROW = ask_box(ws, ROW,
        '요청: 위 3개 항목 ABB 기술 문서(레퍼런스) 또는 담당 전문가 연결 요청')
    ROW += 1

    # 팔로업
    ROW = section('4. 예상 팔로업 항목', ROW)
    ws.row_dimensions[ROW].height = 18
    for col, hdr in [(2, '기한'), (3, '담당'), (4, '내용'), (5, '내용'), (6, '상태')]:
        cell(ws, ROW, col, hdr, bold=True, size=9, bg='2C3E50',
             color='FFFFFF', align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
    ROW += 1
    followups = [
        ('3/6 (금)', 'ABB 김태경', '신규 ADU 삼성중공업 전달 (기약속)', '예정'),
        ('미팅 후 1주', 'ABB', 'ADU 고장 원인 그룹 분석 결과 공유', '대기'),
        ('미팅 후 1주', 'ABB', '권고 스페어 파트 목록 제출', '대기'),
        ('미팅 후 1주', 'ABB', '교육 커리큘럼(안) 제출', '대기'),
        ('4월 8일 이후', '3사 협의', '유지보수 교육 일정 확정 (SP+삼성 인원 포함)', '예정'),
    ]
    for dt, who, what, st in followups:
        ws.row_dimensions[ROW].height = 25
        cell(ws, ROW, 2, dt, size=9, align='center', border=thin())
        cell(ws, ROW, 3, who, size=9, align='center', border=thin())
        ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
        cell(ws, ROW, 4, what, size=9, border=thin(), wrap=True)
        cell(ws, ROW, 6, st, size=9, align='center',
             bg='FADBD8', color='922B21', bold=True, border=thin())
        ROW += 1
    ROW += 1

    # 푸터
    merge(ws, ROW, 2, ROW, 6,
          '작성: 심태양 부장 (SP시스템스)   |   작성일: 2026-03-03   |   문서번호: S25016-AGD-260304',
          size=8, color='888888', italic=True, align='center', border=thin())

    wb.save(_HERE / '260304_ABB_삼성_SP_미팅_아젠다.xlsx')
    print('아젠다 저장 완료')


def make_abb_minutes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ABB미팅회의록"
    a4_setup(ws, landscape=False)

    col_w = [2, 18, 14, 36, 14, 10, 4]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ROW = 1

    # 제목
    ws.row_dimensions[ROW].height = 36
    merge(ws, ROW, 2, ROW, 6,
          'ABB · 삼성중공업 · SP시스템스  기술 협의 회의록',
          bold=True, size=16, color='FFFFFF', bg='1B2631',
          align='center', valign='center', border=medium())
    ROW += 1
    ws.row_dimensions[ROW].height = 15
    merge(ws, ROW, 2, ROW, 6,
          'B라인 ADU 고장 원인 분석 · 긴급 대응 체계 · 유지보수 교육 · IRC5 장기 수급 우려',
          size=9, color='555555', bg='D5D8DC', align='center', border=thin())
    ROW += 1
    ROW += 1

    def section(title, r):
        ws.row_dimensions[r].height = 22
        merge(ws, r, 2, r, 6, title, bold=True, size=11,
              color='FFFFFF', bg='1A3A5C', align='left', border=medium())
        return r + 1

    def row2(ws, r, label, value):
        ws.row_dimensions[r].height = 18
        cell(ws, r, 2, label, bold=True, size=9, bg='D6EAF8', align='center', border=thin())
        merge(ws, r, 3, r, 6, value, size=9, border=thin())
        return r + 1

    # 1. 개요
    ROW = section('1. 회의 개요', ROW)
    ROW = row2(ws, ROW, '일시', '2026년 3월 4일 (수)  14:00')
    ROW = row2(ws, ROW, '장소', '삼성중공업 거제조선소 자동화 실험동')
    ROW = row2(ws, ROW, '안건', 'ADU 고장 원인 분석 · 긴급 대응 체계 · 유지보수 교육 계획 · IRC5 장기 수급 우려')
    ROW = row2(ws, ROW, '작성자', '심태양 부장 (SP시스템스)')
    ROW += 1

    # 2. 참석자
    ROW = section('2. 참석자', ROW)
    ws.row_dimensions[ROW].height = 18
    for col, hdr in [(2, '소속'), (3, '성명'), (4, '직책'), (5, '비고')]:
        cell(ws, ROW, col, hdr, bold=True, size=9, bg='D6EAF8', align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=5, end_row=ROW, end_column=6)
    ROW += 1
    att = [
        ('SP시스템스', '심태양', '부장 / 개발팀', 'DBEAFE', False),
        ('삼성중공업', '이무림', '담당', 'D5F5E3', False),
        ('삼성중공업', '이동혁', '담당', 'D5F5E3', False),
        ('ABB Korea', '김태경', '차장 / Technical Sales', 'FADBD8', False),
        ('ABB Korea', '이영재', '부장 / 서비스팀', 'FADBD8', False),
        ('ABB Korea', '윤건태', '과장 / 서비스팀', 'FADBD8', False),
        ('ABB Korea (불참)', '김유정', '대리 / 서비스팀 — 전주(2/25) 현장 ADU 교체 담당 (본 회의 불참)', 'E8E8E8', True),
    ]
    for so, nm, pos, bg_c, gray in att:
        ws.row_dimensions[ROW].height = 17
        txt_color = 'AAAAAA' if gray else '000000'
        cell(ws, ROW, 2, so, size=9, bg=bg_c, align='center', border=thin(), color=txt_color)
        cell(ws, ROW, 3, nm, size=9, bold=(not gray), bg=bg_c, align='center', border=thin(), color=txt_color)
        ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
        cell(ws, ROW, 4, pos, size=9, bg=bg_c, border=thin(), color=txt_color)
        cell(ws, ROW, 6, '', size=9, bg=bg_c, border=thin())
        ROW += 1
    ROW += 1

    # 3. 논의 내용
    ROW = section('3. 주요 논의 내용', ROW)

    def topic_hdr(ws, r, title, bg='2C3E50'):
        ws.row_dimensions[r].height = 19
        merge(ws, r, 2, r, 6, title, bold=True, size=10,
              color='FFFFFF', bg=bg, border=thin())
        return r + 1

    def point(ws, r, who_bg, text, text_bg='FDFEFE', bold=False, color='000000'):
        ws.row_dimensions[r].height = 28
        cell(ws, r, 2, '▶', size=9, bg=who_bg, align='center', border=thin())
        merge(ws, r, 3, r, 6, text, size=9, bg=text_bg, border=thin(), wrap=True,
              bold=bold, color=color)
        return r + 1

    def warn(ws, r, text):
        ws.row_dimensions[r].height = 32
        cell(ws, r, 2, '⚠', size=9, bg='FADBD8', align='center', border=thin())
        merge(ws, r, 3, r, 6, text, size=9, bg='FDF2F2', color='922B21',
              border=thin(), wrap=True)
        return r + 1

    def note(ws, r, text):
        ws.row_dimensions[r].height = 28
        cell(ws, r, 2, '※', size=9, bg='FDF9E7', align='center', border=thin())
        merge(ws, r, 3, r, 6, text, size=9, bg='FDF9E7', color='784212',
              border=thin(), wrap=True)
        return r + 1

    # 3-1
    ROW = topic_hdr(ws, ROW, '3-1. 상황 공유 — ADU 고장 및 라인 정지 (2026-02-25)')
    ROW = warn(ws, ROW, '2026-02-25(화) B라인 ADU 고장 → 라인 정지 발생. ABB 김유정 대리(서비스팀)가 현장 출동하여 삼성중공업 보유 스페어 ADU로 교체 복구. 라인 정지로 인해 삼성중공업에 심각한 생산 손실이 발생하였음. 현재 ABB 본사에 고장 원인 분석 요청 중 — 결과 미수신 상태.')
    ROW = point(ws, ROW, 'D5F5E3', '[삼성중공업 이무림 프로] 유저 입장에서 원인 규명과 재발 방지 조치가 반드시 필요. 현재 가동 중인 7대 로봇 전체에 동일 문제 발생 가능성 우려.')
    ROW += 1

    # 3-2
    ROW = topic_hdr(ws, ROW, '3-2. ABB 대응 체계 현황')
    ROW = point(ws, ROW, 'FADBD8',
        '[ABB 김태경 차장]\n'
        '• 경남권 담당 서비스 인원: 5명\n'
        '• SLA(대응시간) 픽스 보장 체계 미운영 — 유선 대응 우선, 필요 시 방문\n'
        '• 국내 ABB Korea 정책: 고장 부품 수리(X) → 교체(Replacement) 방식\n'
        '• 원인 분석 결과 수령 전까지 ABB Korea가 할 수 있는 조치 내부 검토 예정')
    ROW = point(ws, ROW, 'D5F5E3',
        '[삼성중공업] 긴급 대응 방법 및 절차에 대해 명확한 기준 제시 요청. 라인 정지 시 ABB 현장 대응까지 소요 시간 확인 요청.')
    ROW += 1

    # 3-3
    ROW = topic_hdr(ws, ROW, '3-3. 유지보수 교육 관련')
    ROW = point(ws, ROW, 'FADBD8',
        '[ABB] 양산·생산 단계에서는 기본 교육 및 유지보수 교육 완료가 원칙. 교육 비용 발생.')
    ROW = note(ws, ROW,
        '이슈 — 교육 비용 입장 차이\n'
        '• 라인 정지 상태에서 현장 인원이 즉시 라인을 재가동할 수 있게 하는 긴급복구 교육 비용은 유저 입장에서 공감·동의 어려운 상황\n'
        '• 이후 진행되는 정식 메인터넌스 교육에 대한 비용 처리는 수긍 가능')
    ROW = point(ws, ROW, 'D5F5E3',
        '[삼성중공업 요구 사항 — 공문 발송 예정, 금주(3/7) 이내 회신 요청]\n'
        '• 원인 분석 회신 기한 명시 요청 — "언제까지" 결과를 줄 것인지 날짜 확정\n'
        '• 긴급 부품 수배 기한 명시 요청 — "몇 일까지" 조치 가능한지 확정\n'
        '• ABB가 제공하는 교육 커리큘럼 내용 및 범위 제시 요청\n'
        '• 교육 핵심 목표: 긴급 상황 발생 시 현장 인원이 즉시 라인을 재가동할 수 있는 능력 확보')
    ROW = note(ws, ROW,
        'ABB 입장 — 기한 확정을 고객사에 공식 제시한 바는 없으나, 유연한 형태로 ABB 내부 협의를 통해 기한을 제시하려는 방향으로 검토 중')
    ROW += 1

    # 3-4
    ROW = topic_hdr(ws, ROW, '3-4. 교육 일정 및 단기 대응 합의')
    ROW = point(ws, ROW, 'DBEAFE',
        '[SP시스템스 심태양 부장]\n'
        '• 4월 8일(녹산 이설 완료)까지 SP시스템스 장비 활용 교육 진행 가능\n'
        '• 지금부터 4/8까지 기간 동안 ABB Korea 최대한 긴급 대응 강력 요청')
    ROW += 1

    # 3-5
    ROW = topic_hdr(ws, ROW, '3-5. IRC5 단종 및 장기 부품 수급 우려', bg='922B21')
    ROW = point(ws, ROW, 'D5D8DC',
        '[삼성중공업 · SP시스템스 공통 우려]\n'
        '• 현재 시스템은 ABB IRC5 컨트롤러 기반 부품 사용 중\n'
        '• 향후 컨트롤러가 OmniCore로 변경될 경우 기존 IRC5 부품 호환성 및 수급 불가 우려\n'
        '• 장비 사용 계획 기간: 향후 20년 — 이 기간 동안 지속적으로 부품 수급 가능 여부에 강한 의문\n'
        '• ABB에 장기 부품 수급 보장 방안 및 로드맵 제시 요청',
        text_bg='FDF2F2', color='922B21')
    ROW = note(ws, ROW,
        '핵심 리스크 — IRC5 단종 + OmniCore 전환 + 20년 운용 계획의 조합은 장기적 유지보수 지속 가능성에 구조적 불확실성을 내포함.\n'
        'ABB의 공식 부품 지원 기간(EOS: End of Service) 및 대체 공급 방안 확인 필수.')
    ROW = warn(ws, ROW,
        '⚠ SP시스템스 강력 확인 요청 — ArcLink XT 호환성 문제\n'
        'OmniCore로 기종 변경 시 Lincoln Electric ArcLink XT 통신 프로토콜이 OmniCore와 호환되지 않는 상황을 이미 전달받은 바 있음.\n'
        '신규 투자(후속 라인)는 Lincoln Electric 용접기와 ArcLink XT 기반으로 통신할 예정이며, IRC5 → OmniCore 전환이 이루어질 경우 용접기 인터페이스 전체가 작동 불가해지는 치명적 문제임.\n'
        '→ ABB는 이 사안의 현재 상황에 대하여 객관적이고 공식적인 피드백을 신속히 제출해야 함')
    ROW += 1

    # 4. 미결 사항
    ROW = section('4. 미결 사항 (Open Items) — ABB 회신 필요', ROW)
    open_items = [
        ('1', 'ADU 고장 원인 분석 결과', 'ABB 본사 분석 완료 후 회신 일정 확정 요청', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('2', '7대 로봇 전체 건전성 검증', 'ABB: 다른 부품에 대한 건전성 평가는 불가하다는 입장 — SP·삼성중공업 대안 별도 검토 필요', 'FDF2F2', '불가', 'FADBD8', '922B21'),
        ('3', '긴급 대응 절차 및 응답 시간', 'ABB 경남권 5인 체계 기준 최대 응답·출동 시간 명확화', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('4', '스페어 파트 권고 수량', '7대 운용 기준 필요 스페어 목록·수량 제시', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('5', '스페어 부품 수급 리드타임 확정', 'ADU 등 주요 부품 긴급 조달 시 소요 기간', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('6', '원인 분석 회신 기한 확정', '"언제까지" 결과를 줄 것인지 날짜 확정', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('7', '교육 커리큘럼 제출', '유지보수 교육 내용 제출 (라인 긴급 재가동 중심)', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('8', 'IRC5→OmniCore 전환 시 부품 호환성', '기존 IRC5 부품 지속 수급 가능 여부 확인', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('9', '20년 장기 부품 수급 보장 방안', 'ABB EOS(End of Service) 정책 및 대체 공급 로드맵 제시 요청', 'FDFEFE', '대기', 'FDEBD0', 'E67E22'),
        ('10', 'OmniCore 전환 시 Lincoln ArcLink XT 호환성 공식 확인', 'IRC5 → OmniCore 변경 시 Lincoln Electric 용접기 인터페이스(ArcLink XT) 작동 가능 여부에 대한 객관적·공식적 피드백 신속 제출 요청', 'FDF2F2', '긴급', 'FADBD8', '922B21'),
    ]
    ws.row_dimensions[ROW].height = 18
    cell(ws, ROW, 2, 'No.', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    cell(ws, ROW, 3, '항목', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
    cell(ws, ROW, 4, '내용', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    cell(ws, ROW, 6, '상태', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    ROW += 1
    for num, item, desc, bg_c, st_txt, st_bg, st_color in open_items:
        ws.row_dimensions[ROW].height = 26
        cell(ws, ROW, 2, num, size=9, bold=True, bg=bg_c, align='center', border=thin())
        cell(ws, ROW, 3, item, size=9, bold=True, bg=bg_c, border=thin())
        ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
        cell(ws, ROW, 4, desc, size=9, bg=bg_c, border=thin(), wrap=True)
        cell(ws, ROW, 6, st_txt, size=9, bold=True, bg=st_bg,
             color=st_color, align='center', border=thin())
        ROW += 1
    ROW += 1

    # 5. 팔로업
    ROW = section('5. 팔로업 액션', ROW)
    ws.row_dimensions[ROW].height = 18
    cell(ws, ROW, 2, '기한', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    cell(ws, ROW, 3, '담당', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
    cell(ws, ROW, 4, '내용', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    cell(ws, ROW, 6, '상태', bold=True, size=9, bg='2C3E50', color='FFFFFF', align='center', border=thin())
    ROW += 1
    followups = [
        ('3/6 (금)',          'ABB 김태경 차장', '신규 ADU 삼성중공업 전달 (기약속)',                                             'FDEBD0', 'E67E22', '예정'),
        ('ABB 내부 협의 후 제시', 'ABB 본사',    'ADU 고장 원인 분석 결과 회신 — 기한 확정 후 통보 예정',                           'FDEBD0', 'E67E22', '대기'),
        ('—',                'ABB',            '7대 로봇 부가축 부품 건전성 평가 — ABB 불가 입장',                               'FADBD8', '922B21', '불가'),
        ('금주(3/7)까지',      'ABB',            '권고 스페어 파트 목록 · 수량 · 수급 리드타임 확정 제출',                         'FDEBD0', 'E67E22', '대기'),
        ('ABB 내부 협의 후 제시', 'ABB',         '긴급 대응 절차 및 최대 응답·출동 시간 기준 제시',                                'FDEBD0', 'E67E22', '대기'),
        ('금주(3/7)까지',      'ABB',            '유지보수 교육 커리큘럼 제출 (라인 긴급 재가동 중심)',                             'FDEBD0', 'E67E22', '대기'),
        ('신속(금주)',          'ABB',            'OmniCore 전환 시 Lincoln ArcLink XT 호환성 — 공식 객관적 피드백 제출\n(IRC5 → OmniCore 변경 시 용접기 인터페이스 작동 불가 우려 — SP 강력 요청)', 'FADBD8', '922B21', '긴급'),
        ('4/8 이후',          '3사 협의',        '유지보수 교육 일정 확정 (SP시스템스 인원 포함)',                                 'D5F5E3', '1D6A39', '예정'),
    ]
    for dt, who, what, st_bg, st_fg, st_txt in followups:
        ws.row_dimensions[ROW].height = 25
        cell(ws, ROW, 2, dt, size=9, align='center', border=thin())
        cell(ws, ROW, 3, who, size=9, align='center', border=thin())
        ws.merge_cells(start_row=ROW, start_column=4, end_row=ROW, end_column=5)
        cell(ws, ROW, 4, what, size=9, border=thin(), wrap=True)
        cell(ws, ROW, 6, st_txt, size=9, bold=True, bg=st_bg,
             color=st_fg, align='center', border=thin())
        ROW += 1
    ROW += 1

    # 6. 종합
    ROW = section('6. 종합 의견', ROW)
    summaries = [
        'ABB 본사 원인 분석 결과 수령 전까지 7대 로봇 동일 고장 가능성이 상존하며, 조기 리스크 해소가 필요',
        'ABB의 SLA(서비스 수준 협약) 미제공 체계는 생산 현장의 긴급 대응에 구조적 한계 — 삼성중공업·SP 모두 명확한 기준 마련 요구',
        'ABB는 다른 부품에 대한 건전성 평가 불가 입장',
        '교육 비용: 라인 긴급복구 중심의 교육 비용은 유저가 공감 어려운 입장이며, 이후 정식 메인터넌스 교육에 대한 비용 처리는 검토 가능',
        'OmniCore 전환 시 Lincoln ArcLink XT 호환 불가 이슈 — 용접기 인터페이스 전체 작동 불능 가능성이 있는 치명적 사안으로, ABB의 공식 피드백을 신속히 수령해야 함',
        'ABB에 요청한 미결 항목(10건) 회신 기한(금주 포함)을 조속히 확정받아야 함',
    ]
    for s in summaries:
        ws.row_dimensions[ROW].height = 20
        cell(ws, ROW, 2, '■', size=9, bg='EAF4FD', align='center', border=thin())
        merge(ws, ROW, 3, ROW, 6, s, size=9, bg='EAF4FD', border=thin())
        ROW += 1
    ROW += 1

    # 푸터
    merge(ws, ROW, 2, ROW, 6,
          '작성: 심태양 부장 (SP시스템스)   |   작성일: 2026-03-04   |   문서번호: S25016-MOM-260304',
          size=8, color='888888', italic=True, align='center', border=thin())

    wb.save('/home/qwe/works/s25016/docs/260304_ABB_삼성_SP_미팅_회의록.xlsx')
    print('ABB 미팅 회의록 저장 완료')


if __name__ == '__main__':
    make_minutes()
    make_agenda()
    make_abb_minutes()
    print('완료')