#!/usr/bin/env python3
"""
전체 거더 상세 분석 (대각거리 방식)
시작점/끝점 변위를 직선거리(유클리드 거리)로 계산
"""

from pathlib import Path
import json
import re
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple
from collections import Counter


class WeldLineValidator_Diagonal:
    """대각거리(직선거리) 방식 검증기"""

    def __init__(self, thresholds):
        self.thresholds = thresholds

    def ccw(self, ax, ay, bx, by, cx, cy):
        return (bx - ax) * (cy - ay) - (by - ay) * (cx - ax)

    def check_crossing(self, s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y):
        d1 = self.ccw(s2x, s2y, e2x, e2y, s1x, s1y)
        d2 = self.ccw(s2x, s2y, e2x, e2y, e1x, e1y)
        d3 = self.ccw(s1x, s1y, e1x, e1y, s2x, s2y)
        d4 = self.ccw(s1x, s1y, e1x, e1y, e2x, e2y)
        return ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
               ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0))

    def check_direction(self, s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y):
        dot = (e1x - s1x) * (e2x - s2x) + (e1y - s1y) * (e2y - s2y)
        return dot > 0

    def get_length(self, sx, sy, ex, ey):
        return math.sqrt((ex - sx) ** 2 + (ey - sy) ** 2)

    def dist_2d(self, x1, y1, x2, y2):
        return math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)

    def point_to_line_dist(self, px, py, x1, y1, x2, y2):
        dx = x2 - x1
        dy = y2 - y1
        if dx == 0 and dy == 0:
            return self.dist_2d(px, py, x1, y1)
        t = max(0, min(1, ((px - x1) * dx + (py - y1) * dy) / (dx * dx + dy * dy)))
        proj_x = x1 + t * dx
        proj_y = y1 + t * dy
        return self.dist_2d(px, py, proj_x, proj_y)

    def check_all(self, line1, line2):
        """모든 검사 수행 (시작점/끝점 변위는 대각거리 사용)"""
        s1x, s1y, s1z = line1['start']['x'], line1['start']['y'], line1['start'].get('z', 0)
        e1x, e1y, e1z = line1['end']['x'], line1['end']['y'], line1['end'].get('z', 0)
        s2x, s2y, s2z = line2['start']['x'], line2['start']['y'], line2['start'].get('z', 0)
        e2x, e2y, e2z = line2['end']['x'], line2['end']['y'], line2['end'].get('z', 0)

        # 1. 교차
        cross = self.check_crossing(s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y)
        cross_ok = not cross

        # 2. 방향
        dir_ok = self.check_direction(s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y)

        # 3. 길이
        len1 = self.get_length(s1x, s1y, e1x, e1y)
        len2 = self.get_length(s2x, s2y, e2x, e2y)
        len_diff = abs(len1 - len2)
        len_ok = len_diff <= self.thresholds['length']

        # 4. 각도
        angle1 = math.atan2(e1y - s1y, e1x - s1x) * 180 / math.pi
        angle2 = math.atan2(e2y - s2y, e2x - s2x) * 180 / math.pi
        angle_diff = abs(angle1 - angle2)
        if angle_diff > 180:
            angle_diff = 360 - angle_diff
        angle_ok = angle_diff <= self.thresholds['angle']

        # 5. 거리 (Line1의 시작/끝점에서 Line2까지)
        dist1 = self.point_to_line_dist(s1x, s1y, s2x, s2y, e2x, e2y)
        dist2 = self.point_to_line_dist(e1x, e1y, s2x, s2y, e2x, e2y)
        dist = max(dist1, dist2)
        dist_ok = dist <= self.thresholds['distance']

        # 6. 시작점 변위 (대각거리 - 직선거리)
        start_off = math.sqrt((s1x - s2x)**2 + (s1y - s2y)**2 + (s1z - s2z)**2)
        start_ok = start_off <= self.thresholds['offset']

        # 7. 끝점 변위 (대각거리 - 직선거리)
        end_off = math.sqrt((e1x - e2x)**2 + (e1y - e2y)**2 + (e1z - e2z)**2)
        end_ok = end_off <= self.thresholds['offset']

        all_ok = cross_ok and dir_ok and len_ok and angle_ok and dist_ok and start_ok and end_ok

        # 실패 사유
        fail_reasons = []
        if not cross_ok:
            fail_reasons.append("교차")
        if not dir_ok:
            fail_reasons.append("방향불일치")
        if not len_ok:
            fail_reasons.append(f"길이차이({len_diff:.1f}mm > {self.thresholds['length']}mm)")
        if not angle_ok:
            fail_reasons.append(f"각도차이({angle_diff:.1f}° > {self.thresholds['angle']}°)")
        if not dist_ok:
            fail_reasons.append(f"거리({dist:.1f}mm > {self.thresholds['distance']}mm)")
        if not start_ok:
            fail_reasons.append(f"시작점변위({start_off:.1f}mm > {self.thresholds['offset']}mm)")
        if not end_ok:
            fail_reasons.append(f"끝점변위({end_off:.1f}mm > {self.thresholds['offset']}mm)")

        return {
            'all_ok': all_ok,
            'cross_ok': cross_ok,
            'dir_ok': dir_ok,
            'len_ok': len_ok,
            'angle_ok': angle_ok,
            'dist_ok': dist_ok,
            'start_ok': start_ok,
            'end_ok': end_ok,
            'len_diff': len_diff,
            'angle_diff': angle_diff,
            'dist': dist,
            'start_off': start_off,
            'end_off': end_off,
            'fail_reasons': ', '.join(fail_reasons) if fail_reasons else 'N/A'
        }


def extract_weld_data(log_file: Path) -> Tuple[List[Dict], Dict]:
    """로그에서 용접선 데이터 및 메타데이터 추출"""
    with open(log_file, 'r', encoding='utf-8') as f:
        content = f.read()

    metadata = {
        'log_file': log_file.name,
        'girder': '알수없음',
        'date': '알수없음'
    }

    parts = log_file.parts
    for part in parts:
        if 'B라인' in part or '거더' in part:
            metadata['girder'] = part
            break

    date_match = re.search(r'(\d{8})', log_file.name)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            metadata['date'] = date_obj.strftime('%Y년 %m월 %d일')
        except:
            metadata['date'] = date_str

    pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ INFO.*?\[GantryRobotControllerServiceImpl\.RequestWeldingOperationStart\(\)\] ENTERED\. (.+?)(?=\n\d{4}-\d{2}-\d{2}|\Z)'
    matches = re.finditer(pattern, content, re.DOTALL)

    weld_data = []
    for match in matches:
        try:
            timestamp = match.group(1)
            json_str = match.group(2).strip()
            data = json.loads(json_str)

            if 'weldJobs' in data and len(data['weldJobs']) >= 2:
                r1 = data['weldJobs'][0]['weldLine']
                r2 = data['weldJobs'][1]['weldLine']

                weld_data.append({
                    'timestamp': timestamp,
                    'requestId': data['request']['requestId'],
                    'line1': {
                        'start': r1['startPointPosition'],
                        'end': r1['endPointPosition']
                    },
                    'line2': {
                        'start': r2['startPointPosition'],
                        'end': r2['endPointPosition']
                    }
                })
        except:
            continue

    return weld_data, metadata


def remove_duplicates(weld_data: List[Dict]) -> Tuple[List[Dict], int]:
    """RequestID 중복 제거"""
    unique_data = []
    seen_ids = set()
    duplicate_count = 0

    for data in weld_data:
        request_id = data.get('requestId')
        if request_id is None:
            unique_data.append(data)
            continue

        if request_id in seen_ids:
            duplicate_count += 1
            continue

        seen_ids.add(request_id)
        unique_data.append(data)

    return unique_data, duplicate_count


def create_summary_sheet(wb, all_results, thresholds):
    """전체 요약 시트"""
    ws = wb.active
    ws.title = '00_전체요약'

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = '용접선 검증 보고서 (대각거리 방식)'
    ws['A1'].font = Font(size=18, bold=True, color='366092')
    ws.merge_cells('A1:I1')

    ws['A3'] = '검증 방식:'
    ws['B3'] = '시작점/끝점 변위를 직선거리(대각거리)로 계산'
    ws['A4'] = '생성 시각:'
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A5'] = '분석 파일 수:'
    ws['B5'] = len(all_results)
    ws['A6'] = '검증 기준:'
    ws['B6'] = f"길이≤{thresholds['length']}mm, 각도≤{thresholds['angle']}°, 거리≤{thresholds['distance']}mm, 변위≤{thresholds['offset']}mm (대각거리)"

    row = 8
    headers = ['거더', '날짜', '총 케이스', 'PASS', 'FAIL', 'PASS율(%)', '고유 FAIL ID', '주요 문제', '시트명']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row += 1
    for result in all_results:
        ws.cell(row, 1, result['girder']).border = border
        ws.cell(row, 2, result['date']).border = border
        ws.cell(row, 3, result['total']).border = border

        pass_cell = ws.cell(row, 4, result['pass'])
        pass_cell.border = border
        if result['pass'] > 0:
            pass_cell.fill = pass_fill

        fail_cell = ws.cell(row, 5, result['fail'])
        fail_cell.border = border
        if result['fail'] > 0:
            fail_cell.fill = fail_fill

        ws.cell(row, 6, f"{result['pass_rate']:.1f}").border = border
        ws.cell(row, 7, result['unique_fail_ids']).border = border
        ws.cell(row, 8, result['main_issue']).border = border
        ws.cell(row, 9, result['sheet_name']).border = border

        row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20


def analyze_single_log(log_file: Path, thresholds: Dict) -> Dict:
    """단일 로그 파일 분석"""
    print(f"  📂 {log_file.parent.name}/{log_file.name}")

    weld_data_raw, metadata = extract_weld_data(log_file)

    if not weld_data_raw:
        return None

    weld_data, duplicate_count = remove_duplicates(weld_data_raw)
    if duplicate_count > 0:
        print(f"     중복 제거: {duplicate_count}개")

    validator = WeldLineValidator_Diagonal(thresholds)
    results = [validator.check_all(d['line1'], d['line2']) for d in weld_data]

    total = len(results)
    passed = sum(1 for r in results if r['all_ok'])
    failed = total - passed

    fail_ids = [weld_data[i]['requestId'] for i, r in enumerate(results) if not r['all_ok']]
    unique_fail_ids = len(set(fail_ids))

    if fail_ids:
        counter = Counter(fail_ids)
        most_common_id, count = counter.most_common(1)[0]
        main_issue = f"ID {most_common_id} ({count}회)"
    else:
        main_issue = "없음"

    girder_short = metadata['girder'].replace('B라인', '').replace('번거더', '')
    date_short = metadata['date'].replace('년 ', '-').replace('월 ', '-').replace('일', '')
    sheet_name = f"{girder_short}_{date_short}"[:31]

    print(f"     ✅ PASS: {passed}, ❌ FAIL: {failed}")

    return {
        'log_file': log_file,
        'metadata': metadata,
        'weld_data': weld_data,
        'results': results,
        'girder': metadata['girder'],
        'date': metadata['date'],
        'total': total,
        'pass': passed,
        'fail': failed,
        'pass_rate': (passed / total * 100) if total > 0 else 0,
        'unique_fail_ids': unique_fail_ids,
        'main_issue': main_issue,
        'sheet_name': sheet_name
    }


def find_all_log_files(base_dir: Path) -> List[Path]:
    """모든 로그 파일 찾기"""
    log_files = []
    for girder_dir in base_dir.glob('B라인*거더'):
        for log_file in girder_dir.rglob('rcs_*.log'):
            if not log_file.name.startswith('rcs_error'):
                log_files.append(log_file)
    return sorted(log_files)


def main():
    base_dir = Path(__file__).parent

    thresholds = {
        'length': 20.0,
        'angle': 5.0,
        'distance': 20.0,
        'offset': 20.0
    }

    print("="*80)
    print("전체 거더 분석 (대각거리 방식)")
    print("="*80)
    print("시작점/끝점 변위: 직선거리(대각거리) 사용")
    print("="*80)

    log_files = find_all_log_files(base_dir)

    if not log_files:
        print("❌ 로그 파일을 찾을 수 없습니다.")
        return

    print(f"\n✅ 총 {len(log_files)}개 로그 파일 발견\n")

    wb = Workbook()
    all_results = []

    for log_file in log_files:
        result = analyze_single_log(log_file, thresholds)
        if result:
            all_results.append(result)

    if not all_results:
        print("\n❌ 분석 가능한 데이터가 없습니다.")
        return

    create_summary_sheet(wb, all_results, thresholds)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = base_dir / f"대각거리방식_7가지검사_{timestamp}.xlsx"
    wb.save(output_file)

    print("\n" + "="*80)
    print("분석 완료!")
    print("="*80)
    print(f"\n✅ 보고서 저장: {output_file}")
    print(f"   - 요약 시트: 1개")


if __name__ == '__main__':
    main()
