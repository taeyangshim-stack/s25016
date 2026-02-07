#!/usr/bin/env python3
"""
용접선 로그 분석 도구 (Excel 버전)
- 거더정보, 날짜시간, 기준값 포함
- 다중 탭 구성 (요약, 전체, FAIL만, 통계)
- 조건부 서식 적용
"""

import json
import math
import re
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class WeldLineAnalyzer:
    """용접선 분석기"""

    def __init__(self, thresholds=None):
        if thresholds is None:
            thresholds = {
                'length': 20.0,
                'angle': 5.0,
                'distance': 20.0,
                'offset': 20.0
            }
        self.thresholds = thresholds

    def ccw(self, ax, ay, bx, by, cx, cy):
        return (bx - ax) * (cy - ay) - (by - ay) * (cx - ax)

    def check_crossing(self, s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y):
        d1 = self.ccw(s2x, s2y, e2x, e2y, s1x, s1y)
        d2 = self.ccw(s2x, s2y, e2x, e2y, e1x, e1y)
        d3 = self.ccw(s1x, s1y, e1x, e1y, s2x, s2y)
        d4 = self.ccw(s1x, s1y, e1x, e1y, e2x, e2y)
        return ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
               ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0))

    def check_direction(self, s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y):
        dot = (e1x - s1x) * (e2x - s2x) + (e1y - s1y) * (e2y - s2y)
        return dot > 0

    def get_length(self, sx, sy, ex, ey):
        return math.sqrt((ex - sx) ** 2 + (ey - sy) ** 2)

    def check_all(self, line1, line2):
        """모든 검사 수행"""
        s1x, s1y = line1['start']['x'], line1['start']['y']
        e1x, e1y = line1['end']['x'], line1['end']['y']
        s2x, s2y = line2['start']['x'], line2['start']['y']
        e2x, e2y = line2['end']['x'], line2['end']['y']

        # 교차
        cross = self.check_crossing(s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y)
        cross_ok = not cross

        # 방향
        dir_ok = self.check_direction(s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y)

        # 길이
        len1 = self.get_length(s1x, s1y, e1x, e1y)
        len2 = self.get_length(s2x, s2y, e2x, e2y)
        len_diff = abs(len1 - len2)
        len_ok = len_diff <= self.thresholds['length']

        # 각도
        angle1 = math.atan2(e1y - s1y, e1x - s1x) * 180 / math.pi
        angle2 = math.atan2(e2y - s2y, e2x - s2x) * 180 / math.pi
        angle_diff = abs(angle1 - angle2)
        if angle_diff > 180:
            angle_diff = 360 - angle_diff
        angle_ok = angle_diff <= self.thresholds['angle']

        # 거리
        mid1x = (s1x + e1x) / 2
        mid1y = (s1y + e1y) / 2
        dx = mid1x - s2x
        dy = mid1y - s2y
        len_sq = (e2x - s2x) ** 2 + (e2y - s2y) ** 2
        if len_sq == 0:
            dist = math.sqrt(dx ** 2 + dy ** 2)
        else:
            t = max(0, min(1, (dx * (e2x - s2x) + dy * (e2y - s2y)) / len_sq))
            proj_x = s2x + t * (e2x - s2x)
            proj_y = s2y + t * (e2y - s2y)
            dist = math.sqrt((mid1x - proj_x) ** 2 + (mid1y - proj_y) ** 2)
        dist_ok = dist <= self.thresholds['distance']

        # 시작점 오프셋 (양방향 투영, XY 평면)
        # S1을 Line2에 투영 → P1, 거리 S2↔P1
        dx2 = e2x - s2x
        dy2 = e2y - s2y
        len_sq2 = dx2 * dx2 + dy2 * dy2
        if len_sq2 == 0:
            p1x, p1y = s2x, s2y
        else:
            t1 = ((s1x - s2x) * dx2 + (s1y - s2y) * dy2) / len_sq2
            t1 = max(0, min(1, t1))
            p1x = s2x + t1 * dx2
            p1y = s2y + t1 * dy2
        dist_s2_p1 = math.sqrt((p1x - s2x) ** 2 + (p1y - s2y) ** 2)

        # S2를 Line1에 투영 → P2, 거리 S1↔P2
        dx1 = e1x - s1x
        dy1 = e1y - s1y
        len_sq1 = dx1 * dx1 + dy1 * dy1
        if len_sq1 == 0:
            p2x, p2y = s1x, s1y
        else:
            t2 = ((s2x - s1x) * dx1 + (s2y - s1y) * dy1) / len_sq1
            t2 = max(0, min(1, t2))
            p2x = s1x + t2 * dx1
            p2y = s1y + t2 * dy1
        dist_s1_p2 = math.sqrt((p2x - s1x) ** 2 + (p2y - s1y) ** 2)

        # 최대값 사용
        start_off = max(dist_s2_p1, dist_s1_p2)
        start_ok = start_off <= self.thresholds['offset']

        # 끝점 오프셋 (양방향 투영, XY 평면)
        # E1을 Line2에 투영 → P3, 거리 E2↔P3
        if len_sq2 == 0:
            p3x, p3y = s2x, s2y
        else:
            t3 = ((e1x - s2x) * dx2 + (e1y - s2y) * dy2) / len_sq2
            t3 = max(0, min(1, t3))
            p3x = s2x + t3 * dx2
            p3y = s2y + t3 * dy2
        dist_e2_p3 = math.sqrt((p3x - e2x) ** 2 + (p3y - e2y) ** 2)

        # E2를 Line1에 투영 → P4, 거리 E1↔P4
        if len_sq1 == 0:
            p4x, p4y = s1x, s1y
        else:
            t4 = ((e2x - s1x) * dx1 + (e2y - s1y) * dy1) / len_sq1
            t4 = max(0, min(1, t4))
            p4x = s1x + t4 * dx1
            p4y = s1y + t4 * dy1
        dist_e1_p4 = math.sqrt((p4x - e1x) ** 2 + (p4y - e1y) ** 2)

        # 최대값 사용
        end_off = max(dist_e2_p3, dist_e1_p4)
        end_ok = end_off <= self.thresholds['offset']

        # 종합
        all_ok = cross_ok and dir_ok and len_ok and angle_ok and dist_ok and start_ok and end_ok

        # 실패 사유
        fail_reasons = []
        if not cross_ok:
            fail_reasons.append('교차')
        if not dir_ok:
            fail_reasons.append('방향')
        if not len_ok:
            fail_reasons.append(f'길이차이({len_diff:.1f}mm > {self.thresholds["length"]}mm)')
        if not angle_ok:
            fail_reasons.append(f'각도차이({angle_diff:.1f}° > {self.thresholds["angle"]}°)')
        if not dist_ok:
            fail_reasons.append(f'거리({dist:.1f}mm > {self.thresholds["distance"]}mm)')
        if not start_ok:
            fail_reasons.append(f'시작점변위({start_off:.1f}mm > {self.thresholds["offset"]}mm)')
        if not end_ok:
            fail_reasons.append(f'끝점변위({end_off:.1f}mm > {self.thresholds["offset"]}mm)')

        return {
            'all_ok': all_ok,
            'cross_ok': cross_ok,
            'dir_ok': dir_ok,
            'len_ok': len_ok,
            'angle_ok': angle_ok,
            'dist_ok': dist_ok,
            'start_ok': start_ok,
            'end_ok': end_ok,
            'len_diff': len_diff,
            'angle_diff': angle_diff,
            'dist': dist,
            'start_off': start_off,
            'end_off': end_off,
            'fail_reasons': ', '.join(fail_reasons) if fail_reasons else 'N/A'
        }


def extract_weld_data(log_file: Path) -> Tuple[List[Dict], Dict]:
    """로그에서 용접선 데이터 및 메타데이터 추출"""
    with open(log_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 메타데이터 추출
    metadata = {
        'log_file': log_file.name,
        'girder': '알수없음',
        'date': '알수없음'
    }

    # 경로에서 거더 정보 추출
    parts = log_file.parts
    for part in parts:
        if 'B라인' in part or '거더' in part:
            metadata['girder'] = part
            break

    # 날짜 추출 (파일명에서)
    date_match = re.search(r'(\d{8})', log_file.name)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            metadata['date'] = date_obj.strftime('%Y년 %m월 %d일')
        except:
            metadata['date'] = date_str

    # 타임스탬프 패턴 매칭
    pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ INFO.*?\[GantryRobotControllerServiceImpl\.RequestWeldingOperationStart\(\)\] ENTERED\. (.+?)(?=\n\d{4}-\d{2}-\d{2}|\Z)'
    matches = re.finditer(pattern, content, re.DOTALL)

    weld_data = []
    for match in matches:
        try:
            timestamp = match.group(1)
            json_str = match.group(2).strip()
            data = json.loads(json_str)

            if 'weldJobs' in data and len(data['weldJobs']) >= 2:
                r1 = data['weldJobs'][0]['weldLine']
                r2 = data['weldJobs'][1]['weldLine']

                weld_data.append({
                    'timestamp': timestamp,
                    'requestId': data['request']['requestId'],
                    'line1': {
                        'start': r1['startPointPosition'],
                        'end': r1['endPointPosition']
                    },
                    'line2': {
                        'start': r2['startPointPosition'],
                        'end': r2['endPointPosition']
                    }
                })
        except:
            continue

    return weld_data, metadata


def remove_duplicates(weld_data: List[Dict]) -> Tuple[List[Dict], int]:
    """RequestID 중복 제거 (첫 번째만 유지)"""
    unique_data = []
    seen_ids = set()
    duplicate_count = 0

    for data in weld_data:
        request_id = data.get('requestId')
        if request_id is None:
            unique_data.append(data)
            continue

        if request_id in seen_ids:
            duplicate_count += 1
            continue

        seen_ids.add(request_id)
        unique_data.append(data)

    return unique_data, duplicate_count


def create_excel_report(results: List[Dict], weld_data: List[Dict],
                        metadata: Dict, thresholds: Dict, output_file: Path):
    """Excel 보고서 생성 (다중 탭)"""
    wb = Workbook()

    # 기본 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 스타일 정의
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. 요약 탭
    ws_summary = wb.create_sheet('요약')
    create_summary_sheet(ws_summary, results, metadata, thresholds, header_fill, header_font)

    # 2. 전체 데이터 탭
    ws_all = wb.create_sheet('전체 데이터')
    create_data_sheet(ws_all, results, weld_data, metadata, thresholds,
                      header_fill, header_font, pass_fill, fail_fill, border, all_data=True)

    # 3. FAIL만 탭
    ws_fail = wb.create_sheet('FAIL만')
    fail_indices = [i for i, r in enumerate(results) if not r['all_ok']]
    fail_results = [results[i] for i in fail_indices]
    fail_data = [weld_data[i] for i in fail_indices]
    create_data_sheet(ws_fail, fail_results, fail_data, metadata, thresholds,
                      header_fill, header_font, pass_fill, fail_fill, border, all_data=False)

    # 4. 통계 탭
    ws_stats = wb.create_sheet('통계')
    create_statistics_sheet(ws_stats, results, metadata, header_fill, header_font)

    wb.save(output_file)
    print(f"\n✅ Excel 저장 완료: {output_file}")
    print(f"   - 탭: 요약, 전체 데이터, FAIL만, 통계")


def create_summary_sheet(ws, results, metadata, thresholds, header_fill, header_font):
    """요약 시트 생성"""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    row = 1

    # 제목
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = '용접선 검증 결과 요약'
    cell.font = Font(size=16, bold=True, color='366092')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # 메타정보
    info_data = [
        ('거더 정보', metadata['girder']),
        ('날짜', metadata['date']),
        ('로그 파일', metadata['log_file']),
        ('생성 시각', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('검사 기준', ''),
        ('  - 길이 오차', f"{thresholds['length']}mm 이하"),
        ('  - 각도 차이', f"{thresholds['angle']}° 이하"),
        ('  - 최대 거리', f"{thresholds['distance']}mm 이하"),
        ('  - 점 변위', f"{thresholds['offset']}mm 이하"),
        ('', ''),
    ]

    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label and not label.startswith('  -'):
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # 검증 결과
    total = len(results)
    passed = sum(1 for r in results if r['all_ok'])
    failed = total - passed

    result_data = [
        ('검증 결과', ''),
        ('총 케이스', f"{total:,}개"),
        ('✅ PASS', f"{passed:,}개 ({passed/total*100:.1f}%)"),
        ('❌ FAIL', f"{failed:,}개 ({failed/total*100:.1f}%)"),
        ('', ''),
    ]

    for label, value in result_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label and not label.startswith('  '):
            ws[f'A{row}'].font = Font(bold=True)
        if 'PASS' in label:
            ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif 'FAIL' in label:
            ws[f'B{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        row += 1

    # 항목별 통계
    ws[f'A{row}'] = '항목별 FAIL 통계'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    fail_stats = {
        '교차': sum(1 for r in results if not r['cross_ok']),
        '방향': sum(1 for r in results if not r['dir_ok']),
        '길이': sum(1 for r in results if not r['len_ok']),
        '각도': sum(1 for r in results if not r['angle_ok']),
        '거리': sum(1 for r in results if not r['dist_ok']),
        '시작점': sum(1 for r in results if not r['start_ok']),
        '끝점': sum(1 for r in results if not r['end_ok'])
    }

    for name, count in fail_stats.items():
        pct = count / total * 100
        ws[f'A{row}'] = f"  {name}"
        ws[f'B{row}'] = f"{count:,}개 ({pct:.1f}%)"
        row += 1


def create_data_sheet(ws, results, weld_data, metadata, thresholds,
                      header_fill, header_font, pass_fill, fail_fill, border, all_data=True):
    """데이터 시트 생성"""
    # 헤더
    headers = [
        '번호', '날짜시간', '거더', 'RequestID', '종합결과', '실패사유',
        '교차', '방향', '길이', '각도', '거리', '시작점', '끝점',
        f'길이차이\n(기준≤{thresholds["length"]}mm)',
        f'각도차이\n(기준≤{thresholds["angle"]}°)',
        f'거리\n(기준≤{thresholds["distance"]}mm)',
        f'시작점변위\n(기준≤{thresholds["offset"]}mm)',
        f'끝점변위\n(기준≤{thresholds["offset"]}mm)',
        'Line1_SX', 'Line1_SY', 'Line1_SZ',
        'Line1_EX', 'Line1_EY', 'Line1_EZ',
        'Line2_SX', 'Line2_SY', 'Line2_SZ',
        'Line2_EX', 'Line2_EY', 'Line2_EZ'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8   # 번호
    ws.column_dimensions['B'].width = 17  # 날짜시간
    ws.column_dimensions['C'].width = 15  # 거더
    ws.column_dimensions['D'].width = 12  # RequestID
    ws.column_dimensions['E'].width = 10  # 결과
    ws.column_dimensions['F'].width = 50  # 실패사유

    for col in range(7, 14):  # 개별 검사 결과
        ws.column_dimensions[get_column_letter(col)].width = 8

    for col in range(14, 19):  # 수치 데이터
        ws.column_dimensions[get_column_letter(col)].width = 12

    for col in range(19, 31):  # 좌표
        ws.column_dimensions[get_column_letter(col)].width = 11

    # 데이터 행
    for idx, (result, data) in enumerate(zip(results, weld_data), 2):
        s1 = data['line1']['start']
        e1 = data['line1']['end']
        s2 = data['line2']['start']
        e2 = data['line2']['end']

        row_data = [
            idx - 1,
            data.get('timestamp', 'N/A'),
            metadata['girder'],
            data['requestId'],
            'PASS' if result['all_ok'] else 'FAIL',
            result['fail_reasons'],
            'OK' if result['cross_ok'] else 'FAIL',
            'OK' if result['dir_ok'] else 'FAIL',
            'OK' if result['len_ok'] else 'FAIL',
            'OK' if result['angle_ok'] else 'FAIL',
            'OK' if result['dist_ok'] else 'FAIL',
            'OK' if result['start_ok'] else 'FAIL',
            'OK' if result['end_ok'] else 'FAIL',
            round(result['len_diff'], 2),
            round(result['angle_diff'], 2),
            round(result['dist'], 2),
            round(result['start_off'], 2),
            round(result['end_off'], 2),
            round(s1['x'], 3), round(s1['y'], 3), round(s1.get('z', 0), 3),
            round(e1['x'], 3), round(e1['y'], 3), round(e1.get('z', 0), 3),
            round(s2['x'], 3), round(s2['y'], 3), round(s2.get('z', 0), 3),
            round(e2['x'], 3), round(e2['y'], 3), round(e2.get('z', 0), 3)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(idx, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 조건부 서식
            if col == 5:  # 종합결과
                if value == 'PASS':
                    cell.fill = pass_fill
                    cell.font = Font(bold=True, color='006100')
                else:
                    cell.fill = fail_fill
                    cell.font = Font(bold=True, color='9C0006')
            elif col >= 7 and col <= 13:  # 개별 검사
                if value == 'FAIL':
                    cell.fill = fail_fill

    # 필터 적용
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 틀 고정 (헤더 행)
    ws.freeze_panes = 'A2'


def create_statistics_sheet(ws, results, metadata, header_fill, header_font):
    """통계 시트 생성"""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    row = 1

    # 제목
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = '상세 통계 분석'
    cell.font = Font(size=14, bold=True, color='366092')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # 수치 통계
    len_diffs = [r['len_diff'] for r in results]
    angle_diffs = [r['angle_diff'] for r in results]
    dists = [r['dist'] for r in results]
    start_offs = [r['start_off'] for r in results]
    end_offs = [r['end_off'] for r in results]

    headers = ['항목', '평균', '최소', '최대', '중앙값']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    stats_data = [
        ('길이 차이 (mm)', len_diffs),
        ('각도 차이 (°)', angle_diffs),
        ('거리 (mm)', dists),
        ('시작점 변위 (mm)', start_offs),
        ('끝점 변위 (mm)', end_offs)
    ]

    for name, values in stats_data:
        sorted_values = sorted(values)
        median = sorted_values[len(sorted_values)//2]

        ws.cell(row, 1, name)
        ws.cell(row, 2, f"{sum(values)/len(values):.2f}")
        ws.cell(row, 3, f"{min(values):.2f}")
        ws.cell(row, 4, f"{max(values):.2f}")
        ws.cell(row, 5, f"{median:.2f}")

        for col in range(1, 6):
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        row += 1


def print_summary(results: List[Dict]):
    """터미널 요약 출력"""
    total = len(results)
    passed = sum(1 for r in results if r['all_ok'])
    failed = total - passed

    print("\n" + "="*70)
    print("📊 검증 결과 요약")
    print("="*70)
    print(f"총 케이스:    {total:,}개")
    print(f"✅ PASS:      {passed:,}개 ({passed/total*100:.1f}%)")
    print(f"❌ FAIL:      {failed:,}개 ({failed/total*100:.1f}%)")


def main():
    """메인 함수"""
    # 로그 파일 경로
    log_file = Path(__file__).parent / "B라인1번거더/02월/04일/rcs_20260204.log"

    print("="*70)
    print("용접선 로그 분석 도구 (Excel 버전)")
    print("="*70)
    print(f"로그 파일: {log_file.name}")

    # 데이터 추출
    print("\n데이터 추출 중...", end='', flush=True)
    weld_data_raw, metadata = extract_weld_data(log_file)
    print(f" 완료! ({len(weld_data_raw):,}개)")

    if not weld_data_raw:
        print("❌ 용접선 데이터를 찾을 수 없습니다.")
        return

    # 중복 제거
    print("중복 제거 중...", end='', flush=True)
    weld_data, duplicate_count = remove_duplicates(weld_data_raw)
    print(f" 완료! (고유: {len(weld_data):,}개, 중복 제거: {duplicate_count:,}개)")

    print(f"거더 정보: {metadata['girder']}")
    print(f"날짜: {metadata['date']}")

    # 검증 수행
    print("\n검증 수행 중...", end='', flush=True)
    thresholds = {
        'length': 20.0,
        'angle': 5.0,
        'distance': 20.0,
        'offset': 20.0
    }
    analyzer = WeldLineAnalyzer(thresholds)
    results = []

    for i, data in enumerate(weld_data):
        if (i + 1) % 500 == 0:
            print(f" {i+1}/{len(weld_data)}", end='', flush=True)
        result = analyzer.check_all(data['line1'], data['line2'])
        results.append(result)

    print(" 완료!")

    # 요약 출력
    print_summary(results)

    # Excel 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file = Path(__file__).parent / f"용접선검증_{metadata['girder']}_{timestamp}.xlsx"

    print("\nExcel 파일 생성 중...", end='', flush=True)
    create_excel_report(results, weld_data, metadata, thresholds, excel_file)

    print("\n" + "="*70)
    print("✅ 분석 완료!")
    print(f"📁 파일: {excel_file.name}")
    print("="*70)


if __name__ == '__main__':
    main()
