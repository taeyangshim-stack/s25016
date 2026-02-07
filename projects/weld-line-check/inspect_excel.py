#!/usr/bin/env python3
"""
Excel 파일 내용 확인 도구
시트 목록, 데이터 샘플, 통계 정보 표시
"""

from pathlib import Path
from openpyxl import load_workbook
import argparse


def inspect_excel(file_path: Path, show_data: bool = False, limit: int = 5):
    """Excel 파일 내용 검사"""

    if not file_path.exists():
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        return

    print("="*80)
    print(f"📊 Excel 파일 검사: {file_path.name}")
    print("="*80)
    print()

    # Excel 파일 로드
    wb = load_workbook(file_path, read_only=True, data_only=True)

    # 전체 시트 목록
    print(f"📑 총 시트 수: {len(wb.sheetnames)}개")
    print()

    # 각 시트 정보
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]

        # 행/열 수
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"{idx}. 📄 시트명: {sheet_name}")
        print(f"   - 크기: {max_row:,}행 × {max_col}열")

        # 헤더 확인
        if max_row > 0:
            headers = []
            for col in range(1, min(max_col + 1, 11)):  # 처음 10개 컬럼만
                cell_value = ws.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value))

            if headers:
                print(f"   - 컬럼: {', '.join(headers[:5])}", end='')
                if len(headers) > 5:
                    print(f" ... 외 {len(headers)-5}개")
                else:
                    print()

        # 데이터 샘플 표시
        if show_data and max_row > 1:
            print(f"   - 데이터 샘플 (처음 {limit}행):")

            # 헤더
            header_row = []
            for col in range(1, min(max_col + 1, 11)):
                header_row.append(str(ws.cell(1, col).value or ''))

            # 데이터 행
            for row_idx in range(2, min(max_row + 1, limit + 2)):
                row_data = []
                for col in range(1, min(max_col + 1, 11)):
                    cell_value = ws.cell(row_idx, col).value
                    if cell_value is None:
                        row_data.append('N/A')
                    elif isinstance(cell_value, (int, float)):
                        row_data.append(f"{cell_value:.2f}" if isinstance(cell_value, float) else str(cell_value))
                    else:
                        row_data.append(str(cell_value)[:20])  # 최대 20자

                print(f"     행{row_idx-1}: {' | '.join(row_data[:5])}")

        print()

    wb.close()

    print("="*80)
    print("💡 팁:")
    print("  - Excel에서 열기: Ctrl+O (Windows) / Cmd+O (Mac)")
    print("  - 시트 간 이동: Ctrl+PgUp/PgDn")
    print("  - 필터 사용: 데이터 탭 → 필터")
    print("="*80)


def list_available_excel_files():
    """사용 가능한 Excel 파일 목록"""
    base_dir = Path(__file__).parent

    excel_files = []
    for pattern in ['*.xlsx', '*.xls']:
        excel_files.extend(base_dir.glob(pattern))

    if not excel_files:
        print("❌ Excel 파일을 찾을 수 없습니다.")
        return

    print("="*80)
    print("📁 사용 가능한 Excel 파일:")
    print("="*80)
    print()

    for idx, file_path in enumerate(sorted(excel_files), 1):
        size_mb = file_path.stat().st_size / 1024 / 1024
        print(f"{idx}. {file_path.name}")
        print(f"   크기: {size_mb:.2f} MB")
        print()


def main():
    parser = argparse.ArgumentParser(
        description='Excel 파일 내용 검사 도구'
    )
    parser.add_argument(
        'file',
        nargs='?',
        type=Path,
        help='검사할 Excel 파일 경로'
    )
    parser.add_argument(
        '--list',
        action='store_true',
        help='사용 가능한 Excel 파일 목록 표시'
    )
    parser.add_argument(
        '--show-data',
        action='store_true',
        help='데이터 샘플 표시'
    )
    parser.add_argument(
        '--limit',
        type=int,
        default=5,
        help='샘플 데이터 행 수 (기본: 5)'
    )

    args = parser.parse_args()

    # 파일 목록 표시
    if args.list:
        list_available_excel_files()
        return

    # 파일 검사
    if args.file:
        inspect_excel(args.file, args.show_data, args.limit)
    else:
        # 기본: 가장 최근 전체거더상세분석 파일
        base_dir = Path(__file__).parent
        excel_files = sorted(
            base_dir.glob('전체거더상세분석_*.xlsx'),
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )

        if excel_files:
            print(f"🔍 가장 최근 파일 자동 선택: {excel_files[0].name}")
            print()
            inspect_excel(excel_files[0], args.show_data, args.limit)
        else:
            print("❌ 전체거더상세분석 파일을 찾을 수 없습니다.")
            print()
            print("사용법:")
            print("  python3 inspect_excel.py <파일명>")
            print("  python3 inspect_excel.py --list")


if __name__ == '__main__':
    main()
