#!/usr/bin/env python3
"""
전체 거더 상세 분석 보고서 생성
모든 거더/날짜별 로그의 상세 데이터를 하나의 Excel 파일에 포함
"""

from pathlib import Path
import json
import re
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple
from collections import Counter


class WeldLineValidator:
    def __init__(self, thresholds):
        self.thresholds = thresholds

    def ccw(self, ax, ay, bx, by, cx, cy):
        return (bx - ax) * (cy - ay) - (by - ay) * (cx - ax)

    def check_crossing(self, s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y):
        d1 = self.ccw(s2x, s2y, e2x, e2y, s1x, s1y)
        d2 = self.ccw(s2x, s2y, e2x, e2y, e1x, e1y)
        d3 = self.ccw(s1x, s1y, e1x, e1y, s2x, s2y)
        d4 = self.ccw(s1x, s1y, e1x, e1y, e2x, e2y)
        return ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
               ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0))

    def check_direction(self, s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y):
        dot = (e1x - s1x) * (e2x - s2x) + (e1y - s1y) * (e2y - s2y)
        return dot > 0

    def get_length(self, sx, sy, ex, ey):
        return math.sqrt((ex - sx) ** 2 + (ey - sy) ** 2)

    def check_all(self, line1, line2):
        """모든 검사 수행"""
        s1x, s1y = line1['start']['x'], line1['start']['y']
        e1x, e1y = line1['end']['x'], line1['end']['y']
        s2x, s2y = line2['start']['x'], line2['start']['y']
        e2x, e2y = line2['end']['x'], line2['end']['y']

        # 교차
        cross = self.check_crossing(s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y)
        cross_ok = not cross

        # 방향
        dir_ok = self.check_direction(s1x, s1y, e1x, e1y, s2x, s2y, e2x, e2y)

        # 길이
        len1 = self.get_length(s1x, s1y, e1x, e1y)
        len2 = self.get_length(s2x, s2y, e2x, e2y)
        len_diff = abs(len1 - len2)
        len_ok = len_diff <= self.thresholds['length']

        # 각도
        angle1 = math.atan2(e1y - s1y, e1x - s1x) * 180 / math.pi
        angle2 = math.atan2(e2y - s2y, e2x - s2x) * 180 / math.pi
        angle_diff = abs(angle1 - angle2)
        if angle_diff > 180:
            angle_diff = 360 - angle_diff
        angle_ok = angle_diff <= self.thresholds['angle']

        # 거리
        dist1 = self.point_to_line_dist(s1x, s1y, s2x, s2y, e2x, e2y)
        dist2 = self.point_to_line_dist(e1x, e1y, s2x, s2y, e2x, e2y)
        dist = max(dist1, dist2)
        dist_ok = dist <= self.thresholds['distance']

        # 시작점 변위 (양방향 투영, XY 평면)
        # S1을 Line2에 투영 → P1, 거리 S2↔P1
        dx2 = e2x - s2x
        dy2 = e2y - s2y
        len_sq2 = dx2 * dx2 + dy2 * dy2
        if len_sq2 == 0:
            p1x, p1y = s2x, s2y
        else:
            t1 = ((s1x - s2x) * dx2 + (s1y - s2y) * dy2) / len_sq2
            t1 = max(0, min(1, t1))
            p1x = s2x + t1 * dx2
            p1y = s2y + t1 * dy2
        dist_s2_p1 = math.sqrt((p1x - s2x) ** 2 + (p1y - s2y) ** 2)

        # S2를 Line1에 투영 → P2, 거리 S1↔P2
        dx1 = e1x - s1x
        dy1 = e1y - s1y
        len_sq1 = dx1 * dx1 + dy1 * dy1
        if len_sq1 == 0:
            p2x, p2y = s1x, s1y
        else:
            t2 = ((s2x - s1x) * dx1 + (s2y - s1y) * dy1) / len_sq1
            t2 = max(0, min(1, t2))
            p2x = s1x + t2 * dx1
            p2y = s1y + t2 * dy1
        dist_s1_p2 = math.sqrt((p2x - s1x) ** 2 + (p2y - s1y) ** 2)

        # 최대값 사용
        start_off = max(dist_s2_p1, dist_s1_p2)
        start_ok = start_off <= self.thresholds['offset']

        # 끝점 변위 (양방향 투영, XY 평면)
        # E1을 Line2에 투영 → P3, 거리 E2↔P3
        if len_sq2 == 0:
            p3x, p3y = s2x, s2y
        else:
            t3 = ((e1x - s2x) * dx2 + (e1y - s2y) * dy2) / len_sq2
            t3 = max(0, min(1, t3))
            p3x = s2x + t3 * dx2
            p3y = s2y + t3 * dy2
        dist_e2_p3 = math.sqrt((p3x - e2x) ** 2 + (p3y - e2y) ** 2)

        # E2를 Line1에 투영 → P4, 거리 E1↔P4
        if len_sq1 == 0:
            p4x, p4y = s1x, s1y
        else:
            t4 = ((e2x - s1x) * dx1 + (e2y - s1y) * dy1) / len_sq1
            t4 = max(0, min(1, t4))
            p4x = s1x + t4 * dx1
            p4y = s1y + t4 * dy1
        dist_e1_p4 = math.sqrt((p4x - e1x) ** 2 + (p4y - e1y) ** 2)

        # 최대값 사용
        end_off = max(dist_e2_p3, dist_e1_p4)
        end_ok = end_off <= self.thresholds['offset']

        all_ok = cross_ok and dir_ok and len_ok and angle_ok and dist_ok and start_ok and end_ok

        # 실패 사유
        fail_reasons = []
        if not cross_ok:
            fail_reasons.append("교차")
        if not dir_ok:
            fail_reasons.append("방향불일치")
        if not len_ok:
            fail_reasons.append(f"길이차이({len_diff:.1f}mm > {self.thresholds['length']}mm)")
        if not angle_ok:
            fail_reasons.append(f"각도차이({angle_diff:.1f}° > {self.thresholds['angle']}°)")
        if not dist_ok:
            fail_reasons.append(f"거리({dist:.1f}mm > {self.thresholds['distance']}mm)")
        if not start_ok:
            fail_reasons.append(f"시작점변위({start_off:.1f}mm > {self.thresholds['offset']}mm)")
        if not end_ok:
            fail_reasons.append(f"끝점변위({end_off:.1f}mm > {self.thresholds['offset']}mm)")

        return {
            'all_ok': all_ok,
            'cross_ok': cross_ok,
            'dir_ok': dir_ok,
            'len_ok': len_ok,
            'angle_ok': angle_ok,
            'dist_ok': dist_ok,
            'start_ok': start_ok,
            'end_ok': end_ok,
            'len_diff': len_diff,
            'angle_diff': angle_diff,
            'dist': dist,
            'start_off': start_off,
            'end_off': end_off,
            'fail_reasons': ', '.join(fail_reasons) if fail_reasons else 'N/A'
        }

    def dist_2d(self, x1, y1, x2, y2):
        return math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)

    def point_to_line_dist(self, px, py, x1, y1, x2, y2):
        dx = x2 - x1
        dy = y2 - y1
        if dx == 0 and dy == 0:
            return self.dist_2d(px, py, x1, y1)
        t = max(0, min(1, ((px - x1) * dx + (py - y1) * dy) / (dx * dx + dy * dy)))
        proj_x = x1 + t * dx
        proj_y = y1 + t * dy
        return self.dist_2d(px, py, proj_x, proj_y)


def extract_weld_data(log_file: Path) -> Tuple[List[Dict], Dict]:
    """로그에서 용접선 데이터 및 메타데이터 추출"""
    with open(log_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 메타데이터 추출
    metadata = {
        'log_file': log_file.name,
        'girder': '알수없음',
        'date': '알수없음'
    }

    # 경로에서 거더 정보 추출
    parts = log_file.parts
    for part in parts:
        if 'B라인' in part or '거더' in part:
            metadata['girder'] = part
            break

    # 날짜 추출 (파일명에서)
    date_match = re.search(r'(\d{8})', log_file.name)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            metadata['date'] = date_obj.strftime('%Y년 %m월 %d일')
        except:
            metadata['date'] = date_str

    # 타임스탬프 패턴 매칭
    pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ INFO.*?\[GantryRobotControllerServiceImpl\.RequestWeldingOperationStart\(\)\] ENTERED\. (.+?)(?=\n\d{4}-\d{2}-\d{2}|\Z)'
    matches = re.finditer(pattern, content, re.DOTALL)

    weld_data = []
    for match in matches:
        try:
            timestamp = match.group(1)
            json_str = match.group(2).strip()
            data = json.loads(json_str)

            if 'weldJobs' in data and len(data['weldJobs']) >= 2:
                r1 = data['weldJobs'][0]['weldLine']
                r2 = data['weldJobs'][1]['weldLine']

                weld_data.append({
                    'timestamp': timestamp,
                    'requestId': data['request']['requestId'],
                    'line1': {
                        'start': r1['startPointPosition'],
                        'end': r1['endPointPosition']
                    },
                    'line2': {
                        'start': r2['startPointPosition'],
                        'end': r2['endPointPosition']
                    }
                })
        except:
            continue

    return weld_data, metadata


def remove_duplicates(weld_data: List[Dict]) -> Tuple[List[Dict], int]:
    """RequestID 중복 제거 (첫 번째만 유지)"""
    unique_data = []
    seen_ids = set()
    duplicate_count = 0

    for data in weld_data:
        request_id = data.get('requestId')
        if request_id is None:
            unique_data.append(data)
            continue

        if request_id in seen_ids:
            duplicate_count += 1
            continue

        seen_ids.add(request_id)
        unique_data.append(data)

    return unique_data, duplicate_count


def create_summary_sheet(wb, all_results, thresholds):
    """전체 요약 시트 생성"""
    ws = wb.active
    ws.title = '00_전체요약'

    # 스타일
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws['A1'] = '용접선 검증 전체 상세 보고서'
    ws['A1'].font = Font(size=18, bold=True, color='366092')
    ws.merge_cells('A1:I1')

    # 생성 정보
    ws['A3'] = '생성 시각:'
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A4'] = '분석 파일 수:'
    ws['B4'] = len(all_results)
    ws['A5'] = '검증 기준:'
    ws['B5'] = f"길이≤{thresholds['length']}mm, 각도≤{thresholds['angle']}°, 거리≤{thresholds['distance']}mm, 변위≤{thresholds['offset']}mm"

    # 헤더
    row = 7
    headers = ['거더', '날짜', '총 케이스', 'PASS', 'FAIL', 'PASS율(%)', '고유 FAIL ID', '주요 문제', '시트명']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 데이터
    row += 1
    for result in all_results:
        ws.cell(row, 1, result['girder']).border = border
        ws.cell(row, 2, result['date']).border = border
        ws.cell(row, 3, result['total']).border = border

        pass_cell = ws.cell(row, 4, result['pass'])
        pass_cell.border = border
        if result['pass'] > 0:
            pass_cell.fill = pass_fill

        fail_cell = ws.cell(row, 5, result['fail'])
        fail_cell.border = border
        if result['fail'] > 0:
            fail_cell.fill = fail_fill

        ws.cell(row, 6, f"{result['pass_rate']:.1f}").border = border
        ws.cell(row, 7, result['unique_fail_ids']).border = border
        ws.cell(row, 8, result['main_issue']).border = border
        ws.cell(row, 9, result['sheet_name']).border = border

        row += 1

    # 열 너비
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20


def create_detail_sheet(wb, sheet_name, results, weld_data, metadata, thresholds):
    """개별 거더/날짜별 상세 데이터 시트 생성"""
    ws = wb.create_sheet(sheet_name)

    # 스타일
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = [
        '번호', '날짜시간', '거더', 'RequestID', '종합결과', '실패사유',
        '교차', '방향', '길이', '각도', '거리', '시작점', '끝점',
        f'길이차이\n(≤{thresholds["length"]}mm)',
        f'각도차이\n(≤{thresholds["angle"]}°)',
        f'거리\n(≤{thresholds["distance"]}mm)',
        f'시작점변위\n(≤{thresholds["offset"]}mm)',
        f'끝점변위\n(≤{thresholds["offset"]}mm)',
        'Line1_SX', 'Line1_SY', 'Line1_SZ',
        'Line1_EX', 'Line1_EY', 'Line1_EZ',
        'Line2_SX', 'Line2_SY', 'Line2_SZ',
        'Line2_EX', 'Line2_EY', 'Line2_EZ'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8   # 번호
    ws.column_dimensions['B'].width = 17  # 날짜시간
    ws.column_dimensions['C'].width = 15  # 거더
    ws.column_dimensions['D'].width = 12  # RequestID
    ws.column_dimensions['E'].width = 10  # 결과
    ws.column_dimensions['F'].width = 50  # 실패사유

    for col in range(7, 14):  # 개별 검사 결과
        ws.column_dimensions[get_column_letter(col)].width = 8

    for col in range(14, 19):  # 수치 데이터
        ws.column_dimensions[get_column_letter(col)].width = 12

    for col in range(19, 31):  # 좌표
        ws.column_dimensions[get_column_letter(col)].width = 11

    # 데이터 행
    for idx, (result, data) in enumerate(zip(results, weld_data), 2):
        s1 = data['line1']['start']
        e1 = data['line1']['end']
        s2 = data['line2']['start']
        e2 = data['line2']['end']

        row_data = [
            idx - 1,
            data.get('timestamp', 'N/A'),
            metadata['girder'],
            data['requestId'],
            'PASS' if result['all_ok'] else 'FAIL',
            result['fail_reasons'],
            'OK' if result['cross_ok'] else 'FAIL',
            'OK' if result['dir_ok'] else 'FAIL',
            'OK' if result['len_ok'] else 'FAIL',
            'OK' if result['angle_ok'] else 'FAIL',
            'OK' if result['dist_ok'] else 'FAIL',
            'OK' if result['start_ok'] else 'FAIL',
            'OK' if result['end_ok'] else 'FAIL',
            round(result['len_diff'], 2),
            round(result['angle_diff'], 2),
            round(result['dist'], 2),
            round(result['start_off'], 2),
            round(result['end_off'], 2),
            round(s1['x'], 3), round(s1['y'], 3), round(s1.get('z', 0), 3),
            round(e1['x'], 3), round(e1['y'], 3), round(e1.get('z', 0), 3),
            round(s2['x'], 3), round(s2['y'], 3), round(s2.get('z', 0), 3),
            round(e2['x'], 3), round(e2['y'], 3), round(e2.get('z', 0), 3)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(idx, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 조건부 서식
            if col == 5:  # 종합결과
                if value == 'PASS':
                    cell.fill = pass_fill
                    cell.font = Font(bold=True, color='006100')
                else:
                    cell.fill = fail_fill
                    cell.font = Font(bold=True, color='9C0006')
            elif col >= 7 and col <= 13:  # 개별 검사
                if value == 'FAIL':
                    cell.fill = fail_fill

    # 필터 적용
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 틀 고정 (헤더 행)
    ws.freeze_panes = 'A2'


def analyze_single_log(log_file: Path, thresholds: Dict) -> Dict:
    """단일 로그 파일 분석"""
    print(f"  📂 {log_file.parent.name}/{log_file.name}")

    weld_data_raw, metadata = extract_weld_data(log_file)

    if not weld_data_raw:
        return None

    # 중복 제거
    weld_data, duplicate_count = remove_duplicates(weld_data_raw)
    if duplicate_count > 0:
        print(f"     중복 제거: {duplicate_count}개 (원본 {len(weld_data_raw)} → 고유 {len(weld_data)})")

    # 검증 수행
    validator = WeldLineValidator(thresholds)
    results = [validator.check_all(d['line1'], d['line2']) for d in weld_data]

    total = len(results)
    passed = sum(1 for r in results if r['all_ok'])
    failed = total - passed

    # 고유 FAIL RequestID 분석
    fail_ids = [weld_data[i]['requestId'] for i, r in enumerate(results) if not r['all_ok']]
    unique_fail_ids = len(set(fail_ids))

    # 주요 문제 분석
    if fail_ids:
        counter = Counter(fail_ids)
        most_common_id, count = counter.most_common(1)[0]
        main_issue = f"ID {most_common_id} ({count}회)"
    else:
        main_issue = "없음"

    # 시트명 생성 (최대 31자)
    girder_short = metadata['girder'].replace('B라인', '').replace('번거더', '')
    date_short = metadata['date'].replace('년 ', '-').replace('월 ', '-').replace('일', '')
    sheet_name = f"{girder_short}_{date_short}"[:31]

    print(f"     ✅ PASS: {passed}, ❌ FAIL: {failed} (고유 FAIL ID: {unique_fail_ids})")

    return {
        'log_file': log_file,
        'metadata': metadata,
        'weld_data': weld_data,
        'results': results,
        'girder': metadata['girder'],
        'date': metadata['date'],
        'total': total,
        'pass': passed,
        'fail': failed,
        'pass_rate': (passed / total * 100) if total > 0 else 0,
        'unique_fail_ids': unique_fail_ids,
        'main_issue': main_issue,
        'sheet_name': sheet_name
    }


def find_all_log_files(base_dir: Path) -> List[Path]:
    """모든 rcs_*.log 파일 찾기"""
    log_files = []

    # B라인1번거더, B라인2번거더, B라인3번거더
    for girder_dir in base_dir.glob('B라인*거더'):
        for log_file in girder_dir.rglob('rcs_*.log'):
            if not log_file.name.startswith('rcs_error'):
                log_files.append(log_file)

    return sorted(log_files)


def main():
    base_dir = Path(__file__).parent

    # 검증 기준
    thresholds = {
        'length': 20.0,
        'angle': 5.0,
        'distance': 20.0,
        'offset': 20.0
    }

    print("="*80)
    print("전체 거더 상세 분석 보고서 생성")
    print("="*80)

    # 로그 파일 검색
    log_files = find_all_log_files(base_dir)

    if not log_files:
        print("❌ 로그 파일을 찾을 수 없습니다.")
        return

    print(f"\n✅ 총 {len(log_files)}개 로그 파일 발견\n")

    # Excel 워크북 생성
    wb = Workbook()

    # 각 로그 파일 분석
    all_results = []
    for log_file in log_files:
        result = analyze_single_log(log_file, thresholds)
        if result:
            all_results.append(result)

            # 상세 시트 생성
            create_detail_sheet(
                wb,
                result['sheet_name'],
                result['results'],
                result['weld_data'],
                result['metadata'],
                thresholds
            )

    if not all_results:
        print("\n❌ 분석 가능한 데이터가 없습니다.")
        return

    # 전체 요약 시트 생성 (맨 앞으로)
    create_summary_sheet(wb, all_results, thresholds)

    # 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = base_dir / f"전체거더상세분석_{timestamp}.xlsx"
    wb.save(output_file)

    print("\n" + "="*80)
    print("분석 완료!")
    print("="*80)
    print(f"\n✅ 상세 보고서 저장: {output_file}")
    print(f"   - 요약 시트: 1개")
    print(f"   - 상세 시트: {len(all_results)}개 (각 거더/날짜별)")
    print(f"   - 전체 시트 수: {len(all_results) + 1}개")


if __name__ == '__main__':
    main()
