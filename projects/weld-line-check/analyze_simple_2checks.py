#!/usr/bin/env python3
"""
대각거리 방식 (2가지 검사) 보고서 생성
시작점/끝점 변위 검사만 수행 (대각거리 측정)
부동소수점 보정: round(distance, 6) < threshold
"""

from pathlib import Path
import json
import re
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple
from collections import Counter


class SimpleValidator:
    """간단 검증기 - 2가지 검사만 수행"""

    ROUND_DIGITS = 6  # 부동소수점 보정 반올림 자릿수

    def __init__(self, offset_threshold):
        self.offset_threshold = offset_threshold

    def check_simple(self, line1, line2):
        """2가지 검사만 수행: 시작점 변위, 끝점 변위 (대각거리 방식)
        부동소수점 보정: round(distance, 6) < threshold
        """
        s1x, s1y = line1['start']['x'], line1['start']['y']
        e1x, e1y = line1['end']['x'], line1['end']['y']
        s2x, s2y = line2['start']['x'], line2['start']['y']
        e2x, e2y = line2['end']['x'], line2['end']['y']

        # Z 좌표 (3D 거리 계산)
        s1z = line1['start'].get('z', 0)
        s2z = line2['start'].get('z', 0)
        e1z = line1['end'].get('z', 0)
        e2z = line2['end'].get('z', 0)

        # 시작점 변위 (대각거리 - 직선거리)
        start_off = math.sqrt((s1x - s2x)**2 + (s1y - s2y)**2 + (s1z - s2z)**2)
        start_off_round = round(start_off, self.ROUND_DIGITS)
        start_ok = start_off_round < self.offset_threshold

        # 끝점 변위 (대각거리 - 직선거리)
        end_off = math.sqrt((e1x - e2x)**2 + (e1y - e2y)**2 + (e1z - e2z)**2)
        end_off_round = round(end_off, self.ROUND_DIGITS)
        end_ok = end_off_round < self.offset_threshold

        all_ok = start_ok and end_ok

        # 실패 사유
        fail_reasons = []
        if not start_ok:
            fail_reasons.append(f"시작점변위({start_off_round}mm >= {self.offset_threshold}mm)")
        if not end_ok:
            fail_reasons.append(f"끝점변위({end_off_round}mm >= {self.offset_threshold}mm)")

        return {
            'all_ok': all_ok,
            'start_ok': start_ok,
            'end_ok': end_ok,
            'start_off': start_off,
            'end_off': end_off,
            'start_off_round': start_off_round,
            'end_off_round': end_off_round,
            'fail_reasons': ', '.join(fail_reasons) if fail_reasons else 'N/A'
        }


def extract_weld_data(log_file: Path) -> Tuple[List[Dict], Dict]:
    """로그에서 용접선 데이터 및 메타데이터 추출"""
    with open(log_file, 'r', encoding='utf-8') as f:
        content = f.read()

    metadata = {
        'log_file': log_file.name,
        'girder': '알수없음',
        'date': '알수없음'
    }

    parts = log_file.parts
    for part in parts:
        if 'B라인' in part or '거더' in part:
            metadata['girder'] = part
            break

    date_match = re.search(r'(\d{8})', log_file.name)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            metadata['date'] = date_obj.strftime('%Y년 %m월 %d일')
        except:
            metadata['date'] = date_str

    pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ INFO.*?\[GantryRobotControllerServiceImpl\.RequestWeldingOperationStart\(\)\] ENTERED\. (.+?)(?=\n\d{4}-\d{2}-\d{2}|\Z)'
    matches = re.finditer(pattern, content, re.DOTALL)

    weld_data = []
    for match in matches:
        try:
            timestamp = match.group(1)
            json_str = match.group(2).strip()
            data = json.loads(json_str)

            if 'weldJobs' in data and len(data['weldJobs']) >= 2:
                r1 = data['weldJobs'][0]['weldLine']
                r2 = data['weldJobs'][1]['weldLine']

                weld_data.append({
                    'timestamp': timestamp,
                    'requestId': data['request']['requestId'],
                    'line1': {
                        'start': r1['startPointPosition'],
                        'end': r1['endPointPosition']
                    },
                    'line2': {
                        'start': r2['startPointPosition'],
                        'end': r2['endPointPosition']
                    }
                })
        except:
            continue

    return weld_data, metadata


def remove_duplicates(weld_data: List[Dict]) -> Tuple[List[Dict], int]:
    """RequestID 중복 제거 (첫 번째만 유지)"""
    unique_data = []
    seen_ids = set()
    duplicate_count = 0

    for data in weld_data:
        request_id = data.get('requestId')
        if request_id is None:
            unique_data.append(data)
            continue

        if request_id in seen_ids:
            duplicate_count += 1
            continue

        seen_ids.add(request_id)
        unique_data.append(data)

    return unique_data, duplicate_count


def create_summary_sheet(wb, all_results, offset_threshold):
    """전체 요약 시트 생성"""
    ws = wb.active
    ws.title = '00_전체요약'

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws['A1'] = '대각거리 방식 (2가지 검사) 보고서 - 부동소수점 보정'
    ws['A1'].font = Font(size=18, bold=True, color='366092')
    ws.merge_cells('A1:I1')

    # 생성 정보
    ws['A3'] = '검증 방식:'
    ws['B3'] = '시작점 변위, 끝점 변위 (대각거리, 2가지 검사만)'
    ws['A4'] = '생성 시각:'
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A5'] = '분석 파일 수:'
    ws['B5'] = len(all_results)
    ws['A6'] = '검증 기준:'
    ws['B6'] = f"round(변위, 6) < {offset_threshold}mm (부동소수점 보정)"
    ws['A7'] = '보정 사유:'
    ws['B7'] = 'sqrt() 부동소수점 오차로 20.0mm 경계값이 19.999... 또는 20.000...으로 산출되는 문제 해결'

    # 헤더
    row = 8
    headers = ['거더', '날짜', '총 케이스', 'PASS', 'FAIL', 'PASS율(%)', '고유 FAIL ID', '주요 문제', '시트명']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 데이터
    row += 1
    for result in all_results:
        ws.cell(row, 1, result['girder']).border = border
        ws.cell(row, 2, result['date']).border = border
        ws.cell(row, 3, result['total']).border = border

        pass_cell = ws.cell(row, 4, result['pass'])
        pass_cell.border = border
        if result['pass'] > 0:
            pass_cell.fill = pass_fill

        fail_cell = ws.cell(row, 5, result['fail'])
        fail_cell.border = border
        if result['fail'] > 0:
            fail_cell.fill = fail_fill

        ws.cell(row, 6, f"{result['pass_rate']:.1f}").border = border
        ws.cell(row, 7, result['unique_fail_ids']).border = border
        ws.cell(row, 8, result['main_issue']).border = border
        ws.cell(row, 9, result['sheet_name']).border = border

        row += 1

    # 열 너비
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20


def create_detail_sheet(wb, sheet_name, results, weld_data, metadata, offset_threshold):
    """개별 거더/날짜별 상세 데이터 시트 생성"""
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = [
        '번호', '날짜시간', '거더', 'RequestID', '종합결과', '실패사유',
        '시작점', '끝점',
        f'시작점변위(raw)',
        f'시작점변위\n(round)',
        f'끝점변위(raw)',
        f'끝점변위\n(round)',
        'Line1_SX', 'Line1_SY', 'Line1_SZ',
        'Line1_EX', 'Line1_EY', 'Line1_EZ',
        'Line2_SX', 'Line2_SY', 'Line2_SZ',
        'Line2_EX', 'Line2_EY', 'Line2_EZ'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 50

    for col in range(7, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14

    for col in range(13, 25):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # 데이터 행
    for idx, (result, data) in enumerate(zip(results, weld_data), 2):
        s1 = data['line1']['start']
        e1 = data['line1']['end']
        s2 = data['line2']['start']
        e2 = data['line2']['end']

        row_data = [
            idx - 1,
            data.get('timestamp', 'N/A'),
            metadata['girder'],
            data['requestId'],
            'PASS' if result['all_ok'] else 'FAIL',
            result['fail_reasons'],
            'OK' if result['start_ok'] else 'FAIL',
            'OK' if result['end_ok'] else 'FAIL',
            result['start_off'],
            result['start_off_round'],
            result['end_off'],
            result['end_off_round'],
            round(s1['x'], 3), round(s1['y'], 3), round(s1.get('z', 0), 3),
            round(e1['x'], 3), round(e1['y'], 3), round(e1.get('z', 0), 3),
            round(s2['x'], 3), round(s2['y'], 3), round(s2.get('z', 0), 3),
            round(e2['x'], 3), round(e2['y'], 3), round(e2.get('z', 0), 3)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(idx, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if col == 5:
                if value == 'PASS':
                    cell.fill = pass_fill
                    cell.font = Font(bold=True, color='006100')
                else:
                    cell.fill = fail_fill
                    cell.font = Font(bold=True, color='9C0006')
            elif col in (7, 8):
                if value == 'FAIL':
                    cell.fill = fail_fill
            elif col in (10, 12):  # round 값 컬럼 - 경계값 강조
                if isinstance(value, (int, float)) and value >= offset_threshold:
                    cell.fill = fail_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = 'A2'


def create_boundary_sheet(wb, all_results, offset_threshold):
    """경계값 분석 시트 생성 (19~21mm 범위 케이스)"""
    ws = wb.create_sheet('경계값분석')

    header_fill = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    changed_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 설명
    ws['A1'] = '경계값 분석 (19.0 ~ 21.0mm 범위)'
    ws['A1'].font = Font(size=14, bold=True, color='8B4513')
    ws.merge_cells('A1:J1')
    ws['A2'] = f'부동소수점 보정 전후 판정 변경 케이스 분석 (threshold={offset_threshold}mm)'
    ws['A2'].font = Font(size=10, color='666666')
    ws.merge_cells('A2:J2')

    # 헤더
    headers = ['거더', 'RequestID', '시작점(raw)', '시작점(round)',
               '끝점(raw)', '끝점(round)', '기존판정\n(<=)', '보정판정\n(round<)', '변경여부']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row = 5
    boundary_count = 0
    changed_count = 0

    for result in all_results:
        for r, d in zip(result['results'], result['weld_data']):
            start_raw = r['start_off']
            end_raw = r['end_off']
            start_round = r['start_off_round']
            end_round = r['end_off_round']

            # 19~21 범위 체크
            in_boundary = ((offset_threshold - 1.0) <= start_raw <= (offset_threshold + 1.0)) or \
                          ((offset_threshold - 1.0) <= end_raw <= (offset_threshold + 1.0))

            if not in_boundary:
                continue

            boundary_count += 1

            # 기존 판정 (<=)
            old_ok = (start_raw <= offset_threshold) and (end_raw <= offset_threshold)
            # 보정 판정 (round <)
            new_ok = (start_round < offset_threshold) and (end_round < offset_threshold)
            changed = old_ok != new_ok

            if changed:
                changed_count += 1

            old_str = 'PASS' if old_ok else 'FAIL'
            new_str = 'PASS' if new_ok else 'FAIL'
            changed_str = 'PASS→FAIL' if (old_ok and not new_ok) else ('FAIL→PASS' if (not old_ok and new_ok) else '-')

            row_data = [
                result['girder'], d['requestId'],
                start_raw, start_round,
                end_raw, end_round,
                old_str, new_str, changed_str
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row, col, value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if col == 7:  # 기존판정
                    cell.fill = pass_fill if value == 'PASS' else fail_fill
                elif col == 8:  # 보정판정
                    cell.fill = pass_fill if value == 'PASS' else fail_fill
                elif col == 9 and changed:  # 변경여부
                    cell.fill = changed_fill
                    cell.font = Font(bold=True, color='9C6500')

            row += 1

    # 요약
    ws.cell(row + 1, 1, f'경계값 케이스: {boundary_count}건').font = Font(bold=True)
    ws.cell(row + 2, 1, f'판정 변경: {changed_count}건').font = Font(bold=True, color='9C6500')

    # 열 너비
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22
    for col in range(7, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = 'A5'


def analyze_single_log(log_file: Path, offset_threshold: float) -> Dict:
    """단일 로그 파일 분석"""
    print(f"  📂 {log_file.parent.name}/{log_file.name}")

    weld_data_raw, metadata = extract_weld_data(log_file)

    if not weld_data_raw:
        return None

    # 중복 제거
    weld_data, duplicate_count = remove_duplicates(weld_data_raw)
    if duplicate_count > 0:
        print(f"     중복 제거: {duplicate_count}개 (원본 {len(weld_data_raw)} → 고유 {len(weld_data)})")

    # 간단 검증 수행
    validator = SimpleValidator(offset_threshold)
    results = [validator.check_simple(d['line1'], d['line2']) for d in weld_data]

    total = len(results)
    passed = sum(1 for r in results if r['all_ok'])
    failed = total - passed

    # 고유 FAIL RequestID 분석
    fail_ids = [weld_data[i]['requestId'] for i, r in enumerate(results) if not r['all_ok']]
    unique_fail_ids = len(set(fail_ids))

    # 주요 문제 분석
    if fail_ids:
        counter = Counter(fail_ids)
        most_common_id, count = counter.most_common(1)[0]
        main_issue = f"ID {most_common_id} ({count}회)"
    else:
        main_issue = "없음"

    # 시트명 생성
    girder_short = metadata['girder'].replace('B라인', '').replace('번거더', '')
    date_short = metadata['date'].replace('년 ', '-').replace('월 ', '-').replace('일', '')
    sheet_name = f"{girder_short}_{date_short}"[:31]

    print(f"     ✅ PASS: {passed}, ❌ FAIL: {failed} (고유 FAIL ID: {unique_fail_ids})")

    return {
        'log_file': log_file,
        'metadata': metadata,
        'weld_data': weld_data,
        'results': results,
        'girder': metadata['girder'],
        'date': metadata['date'],
        'total': total,
        'pass': passed,
        'fail': failed,
        'pass_rate': (passed / total * 100) if total > 0 else 0,
        'unique_fail_ids': unique_fail_ids,
        'main_issue': main_issue,
        'sheet_name': sheet_name
    }


def find_all_log_files(base_dir: Path) -> List[Path]:
    """모든 rcs_*.log 파일 찾기"""
    log_files = []

    for girder_dir in base_dir.glob('B라인*거더'):
        for log_file in girder_dir.rglob('rcs_*.log'):
            if not log_file.name.startswith('rcs_error'):
                log_files.append(log_file)

    return sorted(log_files)


def main():
    base_dir = Path(__file__).parent

    # 검증 기준 (2가지 검사만)
    offset_threshold = 20.0

    print("="*80)
    print("대각거리 방식 (2가지 검사) 보고서 생성 - 부동소수점 보정")
    print("="*80)
    print(f"검증 방식: round(변위, 6) < {offset_threshold}mm")
    print("="*80)

    # 로그 파일 검색
    log_files = find_all_log_files(base_dir)

    if not log_files:
        print("❌ 로그 파일을 찾을 수 없습니다.")
        return

    print(f"\n✅ 총 {len(log_files)}개 로그 파일 발견\n")

    # Excel 워크북 생성
    wb = Workbook()

    # 각 로그 파일 분석
    all_results = []
    for log_file in log_files:
        result = analyze_single_log(log_file, offset_threshold)
        if result:
            all_results.append(result)

            # 상세 시트 생성
            create_detail_sheet(
                wb,
                result['sheet_name'],
                result['results'],
                result['weld_data'],
                result['metadata'],
                offset_threshold
            )

    if not all_results:
        print("\n❌ 분석 가능한 데이터가 없습니다.")
        return

    # 전체 요약 시트 생성
    create_summary_sheet(wb, all_results, offset_threshold)

    # 경계값 분석 시트 생성
    create_boundary_sheet(wb, all_results, offset_threshold)

    # 저장
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = base_dir / f"대각거리방식_2가지검사_반올림보정_{timestamp}.xlsx"
    wb.save(output_file)

    print("\n" + "="*80)
    print("분석 완료!")
    print("="*80)
    print(f"\n✅ 보고서 저장: {output_file}")
    print(f"   - 요약 시트: 1개")
    print(f"   - 상세 시트: {len(all_results)}개 (각 거더/날짜별)")
    print(f"   - 경계값분석 시트: 1개")
    print(f"   - 전체 시트 수: {len(all_results) + 2}개")


if __name__ == '__main__':
    main()
