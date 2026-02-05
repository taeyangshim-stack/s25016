#!/usr/bin/env python3
"""
S25016 펀치리스트 현황 엑셀 생성 스크립트
- 1차 데이터: 최신 CSV (Google Sheet 직접 내보내기)
- 2차 데이터: 3주계획 CSV (비고/작업이력 보완)
- Sheet 1: 미완료 항목 + 2/6~2/8 작업계획 (기능영역별 그룹핑)
- Sheet 2: 요약 (상태별/분류별/담당자별 통계)
- Sheet 3: 전체 목록 (상태별/등록일순 정렬, 이슈 링크 포함)
"""

import csv
import re
from collections import OrderedDict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === Configuration ===
PRIMARY_CSV = "/home/qwe/works/s25016/punchlist/3week-plan/punchlist_2026-02-05 (1).csv"
GANTT_CSV = "/home/qwe/works/s25016/punchlist/3week-plan/S25016_3주완료계획_260130.csv"
OUTPUT_PATH = "/home/qwe/works/s25016/punchlist/3week-plan/S25016_펀치리스트_현황_260205.xlsx"
BASE_URL = "https://s2501602.vercel.app/punchlist/pages/detail.html?id="
PUNCHLIST_URL = "https://s2501602.vercel.app/punchlist/"
GANTT_URL = "https://s2501602.vercel.app/punchlist/3week-plan/"

STATUS_ORDER = {"신규": 0, "진행중": 1, "검증중": 2, "완료": 3, "보류": 4}
STATUS_COLORS = {
    "신규": "BDD7EE", "진행중": "C6EFCE", "검증중": "FFFFCC",
    "완료": "D9D9D9", "보류": "F4CCCC",
}

# === 기능영역 그룹핑 ===
FUNCTIONAL_GROUPS = [
    "알람/팝업",
    "실시간 용접조건",
    "UI 조작 로그",
    "터치에러 핸들링",
    "원점관리",
    "수동티칭",
]

GROUP_COLORS = {
    "알람/팝업":       {"header": "C00000", "row": "F4CCCC"},
    "실시간 용접조건":  {"header": "BF8F00", "row": "FFF2CC"},
    "UI 조작 로그":    {"header": "548235", "row": "E2EFDA"},
    "터치에러 핸들링":  {"header": "4472C4", "row": "D6E4F0"},
    "원점관리":        {"header": "7030A0", "row": "E8D5F5"},
    "수동티칭":        {"header": "C55A11", "row": "FCE4D6"},
    "이벤트성 진행":   {"header": "808080", "row": "F2F2F2"},
}

FUNCTIONAL_AREA_MAP = {
    # 1. 알람/팝업
    "PL-2025-039": "알람/팝업",
    "PL-2025-061": "알람/팝업",
    "PL-2025-067": "알람/팝업",
    "PL-2025-084": "알람/팝업",
    "PL-2026-196": "알람/팝업",
    # 2. 실시간 용접조건
    "PL-2025-065": "실시간 용접조건",
    "PL-2025-091": "실시간 용접조건",
    # 3. UI 조작 로그
    "PL-2025-099": "UI 조작 로그",
    "PL-2025-104": "UI 조작 로그",
    # 4. 터치에러 핸들링
    "PL-2025-100": "터치에러 핸들링",
    "PL-2025-121": "터치에러 핸들링",
    "PL-2025-138": "터치에러 핸들링",
    "PL-2026-178": "터치에러 핸들링",
    "PL-2026-198": "터치에러 핸들링",
    "PL-2026-199": "터치에러 핸들링",
    "PL-2026-208": "터치에러 핸들링",
    "PL-2026-212": "터치에러 핸들링",
    # 5. 원점관리
    "PL-2025-152": "원점관리",
    "PL-2025-153": "원점관리",
    "PL-2025-155": "원점관리",
    "PL-2026-197": "원점관리",
    # 6. 수동티칭
    "PL-2026-171": "수동티칭",
    "PL-2026-173": "수동티칭",
}

PRIORITY_ORDER = {"긴급": 0, "높음": 1, "보통": 2, "낮음": 3, "": 4}

BLOCKER_MAP = {
    "PL-2026-170": "본사지원필요 (Y축 이슈)",
}

# === Styles ===
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
SUBTITLE_FONT = Font(name="맑은 고딕", size=11, bold=True)
LINK_FONT = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

PRIORITY_FILLS = {
    "긴급": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "높음": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
}
PRIORITY_FONTS = {
    "긴급": Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF"),
    "높음": Font(name="맑은 고딕", size=10, bold=True),
}
PLAN_FILLS = {
    "2/6(금)": PatternFill(start_color="FFFFDD", end_color="FFFFDD", fill_type="solid"),
    "2/7(토)": PatternFill(start_color="DDECFF", end_color="DDECFF", fill_type="solid"),
    "2/8(일)": PatternFill(start_color="DDFFE0", end_color="DDFFE0", fill_type="solid"),
}

# 6개 그룹 일정 배분 (Sheet 1: 2/6~2/8)
GROUP_SCHEDULE = {
    "알람/팝업":       {"2/6(금)": "작업예정", "2/7(토)": "디버깅",   "2/8(일)": "테스트"},
    "실시간 용접조건":  {"2/6(금)": "작업예정", "2/7(토)": "테스트",   "2/8(일)": "확인"},
    "UI 조작 로그":    {"2/6(금)": "작업예정", "2/7(토)": "테스트",   "2/8(일)": "확인"},
    "터치에러 핸들링":  {"2/6(금)": "디버깅",   "2/7(토)": "작업예정", "2/8(일)": "테스트"},
    "원점관리":        {"2/6(금)": "-",       "2/7(토)": "작업예정", "2/8(일)": "테스트"},
    "수동티칭":        {"2/6(금)": "-",       "2/7(토)": "작업예정", "2/8(일)": "테스트"},
}

# 6개 그룹 일정 배분 (Sheet 4: 2/5~2/8)
GROUP_SCHEDULE_DAILY = {
    "알람/팝업":       {"2/5(목)": "분석",   "2/6(금)": "작업예정", "2/7(토)": "디버깅",   "2/8(일)": "테스트"},
    "실시간 용접조건":  {"2/5(목)": "-",     "2/6(금)": "작업예정", "2/7(토)": "테스트",   "2/8(일)": "확인"},
    "UI 조작 로그":    {"2/5(목)": "-",     "2/6(금)": "작업예정", "2/7(토)": "테스트",   "2/8(일)": "확인"},
    "터치에러 핸들링":  {"2/5(목)": "분석",   "2/6(금)": "디버깅",   "2/7(토)": "작업예정", "2/8(일)": "테스트"},
    "원점관리":        {"2/5(목)": "-",     "2/6(금)": "-",       "2/7(토)": "작업예정", "2/8(일)": "테스트"},
    "수동티칭":        {"2/5(목)": "-",     "2/6(금)": "-",       "2/7(토)": "작업예정", "2/8(일)": "테스트"},
}


def style(cell, font=None, fill=None, align=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border


def parse_date(iso_str):
    """Parse ISO date string to formatted date."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except:
        return str(iso_str)[:10]


def parse_pl_number(pl_id):
    """Parse PL ID for sorting."""
    m = re.match(r'PL-(\d{4})-(\d{3})', pl_id)
    return (int(m.group(1)), int(m.group(2))) if m else (9999, 9999)


def load_csv_overrides():
    """Load 3-week plan CSV to get 비고/작업이력 overrides."""
    overrides = {}
    with open(GANTT_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if not row or len(row) < 5:
                continue
            if row[0].startswith("===") or not row[0].strip():
                continue
            pl_id = row[0].strip()
            if not pl_id.startswith("PL-"):
                continue
            bigo = row[-1].strip() if len(row) > 4 else ""
            status = extract_status_from_bigo(bigo)
            if not status:
                date_vals = [row[i].strip() for i in range(4, len(row) - 1)]
                status = extract_status_from_dates(date_vals, bigo)
            if status and "→" in bigo:
                after = bigo.split("→")[-1].strip()
                if "보류" in after:
                    status = "보류"
                elif "완료" in after:
                    status = "완료"
                elif "진행" in after:
                    status = "진행중"
            if status:
                overrides[pl_id] = {"status": status, "bigo": bigo}
    return overrides


def extract_status_from_bigo(bigo):
    """Extract definitive status from 비고 field."""
    if not bigo:
        return None
    entries = re.findall(r'\[([\d/]+)\]\s*([^\[]+)', bigo)
    if entries:
        last_text = entries[-1][1].strip()
        if "→" in last_text:
            after_arrow = last_text.split("→")[-1].strip()
            if "보류" in after_arrow:
                return "보류"
            if "완료" in after_arrow:
                return "완료"
            if "진행" in after_arrow:
                return "진행중"
        first_word = last_text.split()[0].split('/')[0] if last_text else ""
        if "완료" in first_word:
            return "완료"
        if "보류" in first_word:
            return "보류"
        if "진행중" in first_word:
            return "진행중"
        if "신규등록" in first_word or first_word == "신규":
            return "신규"
    return None


def extract_status_from_dates(date_vals, bigo):
    """Extract status from date column values."""
    latest = ""
    for v in reversed(date_vals):
        if v and v != "-" and v != "휴무":
            latest = v
            break
    if not latest:
        return None

    direct = {"완료": "완료", "보류": "보류", "진행중": "진행중", "검증중": "검증중", "신규": "신규"}
    if latest in direct:
        return direct[latest]

    activity = {
        "개발", "디버깅", "분석", "UI이슈", "UI개발", "코드준비", "코드완료",
        "측정", "조치예정", "진행", "진행예정", "이슈발생", "삼성검토중",
        "피드백대기", "컨펌대기", "UI보완대기", "배포대기", "예외수정",
        "PlanC연기", "미진행", "A라인완료", "B라인적용", "CSV작성",
        "삼성발송", "확인완료", "품질확인완료", "관제연동", "UI준비",
        "확인", "지연",
    }
    if latest in activity:
        return "진행중"
    if latest in ("테스트완료", "테스트"):
        if any(k in bigo for k in ("이슈", "미반영", "개선필요")):
            return "진행중"
        return "완료"
    return None


def extract_related_from_bigo(bigo, current_id):
    """Extract related PL issue IDs from 비고."""
    if not bigo:
        return ""
    matches = re.findall(r'PL-\d{4}-\d{3}', bigo)
    related = [m for m in set(matches) if m != current_id]
    return ", ".join(sorted(related))


def extract_hold_reason(bigo, status):
    """Extract hold reason for 보류 items."""
    if status != "보류" or not bigo:
        return ""
    if "중복" in bigo:
        dup = re.search(r'중복→?(PL-\d{4}-\d{3})', bigo)
        return f"중복 ({dup.group(1)})" if dup else "중복 항목"
    return ""


def load_primary_data():
    """Load full punchlist from latest CSV export."""
    items = []
    with open(PRIMARY_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            items.append({
                'id': row['ID'].strip(),
                'title': row['제목'].strip(),
                'line': row.get('라인', '-').strip() or '-',
                'category': row.get('분류', '').strip(),
                'subcategory': row.get('세부분류', '').strip(),
                'priority': row.get('우선순위', '').strip(),
                'status': row.get('상태', '신규').strip(),
                'owner': row.get('담당자', '').strip(),
                'request_date': parse_date(row.get('요청일', '')),
                'target_date': parse_date(row.get('목표일', '')),
                'last_changed': parse_date(row.get('최근 변경', '')),
            })
    return items


def build_items():
    """Build complete item list from primary CSV + gantt CSV overrides."""
    primary = load_primary_data()
    csv_overrides = load_csv_overrides()

    items = []
    override_count = 0

    for raw in primary:
        pl_id = raw['id']
        status = raw['status']

        bigo = ""
        related = ""
        hold_reason = ""

        if pl_id in csv_overrides:
            override = csv_overrides[pl_id]
            csv_status = override["status"]
            bigo = override["bigo"]
            related = extract_related_from_bigo(bigo, pl_id)

            if csv_status and csv_status != status:
                override_count += 1
                status = csv_status

            hold_reason = extract_hold_reason(bigo, status)

        complete_date = ""
        if status == "완료" and raw['last_changed']:
            complete_date = raw['last_changed']

        items.append({
            "id": pl_id,
            "title": raw['title'],
            "line": raw['line'],
            "category": raw['category'],
            "subcategory": raw['subcategory'],
            "owner": raw['owner'],
            "status": status,
            "priority": raw['priority'],
            "request_date": raw['request_date'],
            "target_date": raw['target_date'],
            "complete_date": complete_date,
            "description": "",
            "hold_reason": hold_reason,
            "related": related,
            "bigo": bigo,
            "link": f"{BASE_URL}{pl_id}" if pl_id.startswith("PL-") else "",
            "sort_key": parse_pl_number(pl_id),
        })

    print(f"  최신 CSV 항목: {len(primary)}건")
    print(f"  간트차트 오버라이드 적용: {override_count}건")
    return items


def sort_items(items):
    return sorted(items, key=lambda x: (
        STATUS_ORDER.get(x["status"], 99),
        x["sort_key"][0],
        x["sort_key"][1],
    ))


# === Sheet 1: 미완료 항목 + 계획 ===

def filter_incomplete(items):
    """Filter to items that are not 완료 or 보류."""
    return [i for i in items if i["status"] not in ("완료", "보류")]


def classify_functional_area(item):
    """Classify item into functional area group."""
    pl_id = item["id"]
    if pl_id in FUNCTIONAL_AREA_MAP:
        return FUNCTIONAL_AREA_MAP[pl_id]
    return "이벤트성 진행"


def group_and_sort_incomplete(incomplete_items):
    """Group incomplete items by functional area, sort within groups."""
    grouped = OrderedDict()
    for g in FUNCTIONAL_GROUPS:
        grouped[g] = []
    grouped["이벤트성 진행"] = []

    for item in incomplete_items:
        area = classify_functional_area(item)
        if area in grouped:
            grouped[area].append(item)
        else:
            grouped["이벤트성 진행"].append(item)

    for g in grouped:
        grouped[g].sort(key=lambda x: (
            PRIORITY_ORDER.get(x["priority"], 99),
            x["sort_key"][0],
            x["sort_key"][1],
        ))

    return OrderedDict((k, v) for k, v in grouped.items() if v)


def get_blocker_info(item):
    """Get dependency/blocker information for an item."""
    pl_id = item["id"]
    if pl_id in BLOCKER_MAP:
        return BLOCKER_MAP[pl_id]

    bigo = item.get("bigo", "")
    if not bigo:
        return ""

    blockers = []
    if "본사" in bigo and "지원" in bigo:
        blockers.append("본사지원필요")
    elif "본사" in bigo:
        blockers.append("본사 확인")
    if "대기" in bigo and "본사" not in bigo:
        blockers.append("대기중")
    if "삼성" in bigo and "완료" not in bigo:
        blockers.append("삼성 확인")

    return ", ".join(blockers)


def create_plan_sheet(ws, grouped_items):
    """Create Sheet 1: 미완료 항목 + 2/6~2/8 작업 계획"""
    ws.title = "미완료 항목 + 계획"

    total_incomplete = sum(len(v) for v in grouped_items.values())

    # Title rows
    headers = [
        ("No.", 5), ("기능영역", 16), ("ID", 16), ("항목명", 42),
        ("라인", 8), ("담당자", 10), ("우선순위", 8), ("현재상태", 10),
        ("2/6(금)", 14), ("2/7(토)", 14), ("2/8(일)", 14),
        ("의존성/블로커", 20), ("비고", 30), ("이슈 링크", 50),
    ]
    last_col = get_column_letter(len(headers))

    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = "S25016 미완료 항목 현황 및 2/6~2/8 작업 계획"
    c.font = TITLE_FONT
    c.alignment = CENTER

    # Status counts
    scnt = {}
    for items in grouped_items.values():
        for i in items:
            scnt[i["status"]] = scnt.get(i["status"], 0) + 1
    status_str = ", ".join(f"{s} {scnt.get(s, 0)}" for s in ["신규", "진행중", "검증중"] if scnt.get(s, 0))

    ws.merge_cells(f'A2:{last_col}2')
    today = datetime.now().strftime('%Y년 %m월 %d일')
    ws['A2'].value = f"{today} 기준 | 미완료 {total_incomplete}건 ({status_str})"
    ws['A2'].font = NORMAL_FONT
    ws['A2'].alignment = CENTER

    # Owner summary
    ocnt = {}
    for items in grouped_items.values():
        for i in items:
            o = i["owner"] or "미배정"
            ocnt[o] = ocnt.get(o, 0) + 1
    owner_str = ", ".join(f"{k} {v}건" for k, v in sorted(ocnt.items(), key=lambda x: -x[1]))

    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'].value = f"담당자: {owner_str}"
    ws['A3'].font = NORMAL_FONT
    ws['A3'].alignment = CENTER

    # Vercel links (row 4)
    ws.merge_cells(f'A4:G4')
    link_cell = ws['A4']
    link_cell.value = f"펀치리스트: {PUNCHLIST_URL}"
    link_cell.hyperlink = PUNCHLIST_URL
    link_cell.font = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
    link_cell.alignment = LEFT

    ws.merge_cells(f'H4:{last_col}4')
    gantt_cell = ws['H4']
    gantt_cell.value = f"3주계획/간트차트: {GANTT_URL}"
    gantt_cell.hyperlink = GANTT_URL
    gantt_cell.font = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
    gantt_cell.alignment = LEFT

    # Header row (row 5)
    ROW_HEADER = 5
    for col, (name, width) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
        cell = ws.cell(row=ROW_HEADER, column=col, value=name)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    ws.freeze_panes = 'A6'
    ws.auto_filter.ref = f"A{ROW_HEADER}:{last_col}{ROW_HEADER}"

    # Data rows
    row = ROW_HEADER + 1
    seq = 0

    # 6개 사용자 그룹 → 이벤트성 진행 순서
    all_group_names = list(FUNCTIONAL_GROUPS) + (["이벤트성 진행"] if "이벤트성 진행" in grouped_items else [])

    for group_name in all_group_names:
        if group_name not in grouped_items or not grouped_items[group_name]:
            continue

        group_items = grouped_items[group_name]
        colors = GROUP_COLORS.get(group_name, GROUP_COLORS["이벤트성 진행"])
        is_user_group = group_name in FUNCTIONAL_GROUPS

        # Group separator row
        gfill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1,
                       value=f"▶ {group_name} ({len(group_items)}건)")
        style(cell, Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF"),
              gfill, LEFT, THIN_BORDER)
        ws.row_dimensions[row].height = 25
        row += 1

        row_tint = PatternFill(start_color=colors["row"], end_color=colors["row"], fill_type="solid")
        first_data_row = row

        # Pre-filled schedule
        schedule = GROUP_SCHEDULE.get(group_name, {})

        for item in group_items:
            seq += 1
            blocker = get_blocker_info(item)

            # Plan columns: grouped → schedule, non-grouped → 이벤트성
            if is_user_group:
                plan_6 = schedule.get("2/6(금)", "")
                plan_7 = schedule.get("2/7(토)", "")
                plan_8 = schedule.get("2/8(일)", "")
            else:
                plan_6 = "이벤트성"
                plan_7 = "이벤트성"
                plan_8 = "이벤트성"

            area_display = group_name if is_user_group else "-"

            data = [
                seq, area_display, item["id"], item["title"],
                item["line"], item["owner"], item["priority"], item["status"],
                plan_6, plan_7, plan_8,
                blocker, item["bigo"], "",
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)

                align = CENTER if col in (1, 2, 5, 6, 7, 8) else LEFT
                font = NORMAL_FONT
                fill = row_tint

                # Priority column
                if col == 7:
                    p = item["priority"]
                    if p in PRIORITY_FILLS:
                        fill = PRIORITY_FILLS[p]
                    if p in PRIORITY_FONTS:
                        font = PRIORITY_FONTS[p]

                # Status column
                if col == 8:
                    sfill = PatternFill(
                        start_color=STATUS_COLORS.get(item["status"], "FFFFFF"),
                        end_color=STATUS_COLORS.get(item["status"], "FFFFFF"),
                        fill_type="solid")
                    fill = sfill
                    font = BOLD_FONT

                # Plan columns
                if col == 9:
                    fill = PLAN_FILLS["2/6(금)"]
                elif col == 10:
                    fill = PLAN_FILLS["2/7(토)"]
                elif col == 11:
                    fill = PLAN_FILLS["2/8(일)"]

                style(cell, font, fill, align, THIN_BORDER)

            # Hyperlink
            if item["link"]:
                lc = ws.cell(row=row, column=14, value=item["link"])
                lc.hyperlink = item["link"]
                lc.font = LINK_FONT
                lc.alignment = LEFT
                lc.border = THIN_BORDER

            ws.row_dimensions[row].height = 28
            row += 1

        # 사용자 그룹: 기능영역 컬럼(B) 셀 병합으로 그룹 표시
        if is_user_group and len(group_items) > 1:
            ws.merge_cells(f'B{first_data_row}:B{row - 1}')
            mc = ws.cell(row=first_data_row, column=2)
            mc.value = group_name
            mc.font = Font(name="맑은 고딕", size=10, bold=True)
            mc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            mc.fill = row_tint
            mc.border = THIN_BORDER

    # Data Validation for plan columns
    dv = DataValidation(
        type="list",
        formula1='"작업예정,디버깅,테스트,배포,확인,분석,이벤트성,미진행,-"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "목록에서 선택하세요"
    dv.errorTitle = "입력 오류"
    ws.add_data_validation(dv)
    dv.add(f"I{ROW_HEADER + 1}:K{row - 1}")

    # Footer
    row += 1
    ws.merge_cells(f'A{row}:{last_col}{row}')
    ws[f'A{row}'].value = f"총 미완료 {total_incomplete}건 | 2/6~2/8 열에 작업계획을 입력하세요"
    ws[f'A{row}'].font = Font(name="맑은 고딕", size=10, italic=True, color="666666")
    ws[f'A{row}'].alignment = LEFT

    # Print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


# === Sheet 2: 요약 ===

def create_summary_sheet(ws, items):
    ws.title = "요약"
    total = len(items)

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = "S25016 펀치리스트 현황 요약"
    c.font = TITLE_FONT
    c.alignment = CENTER

    ws.merge_cells('A2:F2')
    today = datetime.now().strftime('%Y년 %m월 %d일')
    ws['A2'].value = f"작성일: {today} | 총 {total}건"
    ws['A2'].font = NORMAL_FONT
    ws['A2'].alignment = CENTER

    ws.merge_cells('A3:F3')
    lc = ws['A3']
    lc.value = f"펀치리스트: {PUNCHLIST_URL}"
    lc.hyperlink = PUNCHLIST_URL
    lc.font = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
    lc.alignment = CENTER

    # 상태별
    row = 5
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "■ 상태별 현황"
    ws[f'A{row}'].font = SUBTITLE_FONT

    row += 1
    for col, h in enumerate(["상태", "건수", "비율"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    scnt = {}
    for i in items:
        scnt[i["status"]] = scnt.get(i["status"], 0) + 1

    for sn in ["신규", "진행중", "검증중", "완료", "보류"]:
        cnt = scnt.get(sn, 0)
        row += 1
        fill = PatternFill(start_color=STATUS_COLORS[sn], end_color=STATUS_COLORS[sn], fill_type="solid")
        for col, val in enumerate([sn, cnt, f"{cnt/total*100:.1f}%"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            style(cell, BOLD_FONT, fill, CENTER, THIN_BORDER)

    row += 1
    tfill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, val in enumerate(["합계", total, "100%"], 1):
        cell = ws.cell(row=row, column=col, value=val)
        style(cell, BOLD_FONT, tfill, CENTER, THIN_BORDER)

    # 분류별
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "■ 분류별 현황"
    ws[f'A{row}'].font = SUBTITLE_FONT

    row += 1
    for col, h in enumerate(["분류", "건수", "비율"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    ccnt = {}
    for i in items:
        c = i["category"] or "미분류"
        ccnt[c] = ccnt.get(c, 0) + 1
    for cn in sorted(ccnt):
        cnt = ccnt[cn]
        row += 1
        for col, val in enumerate([cn, cnt, f"{cnt/total*100:.1f}%"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            style(cell, NORMAL_FONT, None, CENTER, THIN_BORDER)

    # 라인별
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "■ 라인별 현황"
    ws[f'A{row}'].font = SUBTITLE_FONT

    row += 1
    for col, h in enumerate(["라인", "건수", "비율"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    lcnt = {}
    for i in items:
        ln = i["line"] or "-"
        lcnt[ln] = lcnt.get(ln, 0) + 1
    for ln in sorted(lcnt):
        cnt = lcnt[ln]
        row += 1
        for col, val in enumerate([ln, cnt, f"{cnt/total*100:.1f}%"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            style(cell, NORMAL_FONT, None, CENTER, THIN_BORDER)

    # 담당자별
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].value = "■ 담당자별 현황"
    ws[f'A{row}'].font = SUBTITLE_FONT

    row += 1
    for col, h in enumerate(["담당자", "전체", "신규", "진행중", "완료", "보류"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    ps = {}
    for i in items:
        p = i["owner"] or "미배정"
        if p not in ps:
            ps[p] = {"전체": 0, "신규": 0, "진행중": 0, "검증중": 0, "완료": 0, "보류": 0}
        ps[p]["전체"] += 1
        ps[p][i["status"]] += 1

    for person in sorted(ps, key=lambda x: ps[x]["전체"], reverse=True):
        s = ps[person]
        row += 1
        for col, val in enumerate([person, s["전체"], s["신규"], s["진행중"] + s["검증중"], s["완료"], s["보류"]], 1):
            cell = ws.cell(row=row, column=col, value=val)
            style(cell, NORMAL_FONT, None, CENTER, THIN_BORDER)

    for c, w in [('A', 15), ('B', 10), ('C', 10), ('D', 10), ('E', 10), ('F', 10)]:
        ws.column_dimensions[c].width = w


# === Sheet 3: 전체 목록 ===

def create_list_sheet(ws, items):
    ws.title = "전체 목록"

    headers = [
        ("No.", 5), ("ID", 16), ("기능영역", 16), ("항목명", 40), ("라인", 10),
        ("분류", 8), ("담당자", 10), ("현재 상태", 10), ("우선순위", 8),
        ("등록일", 12), ("목표일", 12), ("완료일", 12),
        ("보류 사유", 18), ("연관 이슈", 18), ("비고(3주계획)", 45),
        ("이슈 링크", 55),
    ]

    for col, (_, w) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for col, (name, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    current_group = None

    for idx, item in enumerate(items, 1):
        if item["status"] != current_group:
            current_group = item["status"]
            cnt = sum(1 for i in items if i["status"] == current_group)
            gfill = PatternFill(start_color=STATUS_COLORS.get(current_group, "FFFFFF"),
                                end_color=STATUS_COLORS.get(current_group, "FFFFFF"), fill_type="solid")
            ws.merge_cells(f'A{row}:{get_column_letter(len(headers))}{row}')
            cell = ws.cell(row=row, column=1, value=f"▶ {current_group} ({cnt}건)")
            style(cell, Font(name="맑은 고딕", size=11, bold=True), gfill, LEFT, THIN_BORDER)
            ws.row_dimensions[row].height = 25
            row += 1

        sfill = PatternFill(start_color=STATUS_COLORS.get(item["status"], "FFFFFF"),
                            end_color=STATUS_COLORS.get(item["status"], "FFFFFF"), fill_type="solid")

        area = classify_functional_area(item)
        area_display = area if item["id"] in FUNCTIONAL_AREA_MAP else ""

        data = [
            idx, item["id"], area_display, item["title"], item["line"],
            item["category"], item["owner"], item["status"], item["priority"],
            item["request_date"], item["target_date"], item["complete_date"],
            item["hold_reason"], item["related"], item["bigo"], ""
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            align = CENTER if col in (1, 3, 5, 6, 7, 8, 9, 10, 11, 12) else LEFT
            font = BOLD_FONT if col == 8 else NORMAL_FONT
            fill = sfill if col == 8 else None
            # 기능영역 컬럼 색상 (col 3)
            if col == 3 and area_display and area_display in GROUP_COLORS:
                gc = GROUP_COLORS[area_display]
                fill = PatternFill(start_color=gc["row"], end_color=gc["row"], fill_type="solid")
                font = Font(name="맑은 고딕", size=9, bold=True)
            style(cell, font, fill, align, THIN_BORDER)

        if item["link"]:
            lc = ws.cell(row=row, column=len(headers), value=item["link"])
            lc.hyperlink = item["link"]
            lc.font = LINK_FONT
            lc.alignment = LEFT
            lc.border = THIN_BORDER

        ws.row_dimensions[row].height = 22
        row += 1


# === Sheet 4: 일일계획 ===

PLAN_DAYS = ["2/5(목)", "2/6(금)", "2/7(토)", "2/8(일)"]
PLAN_DAY_FILLS = {
    "2/5(목)": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "2/6(금)": PatternFill(start_color="FFFFDD", end_color="FFFFDD", fill_type="solid"),
    "2/7(토)": PatternFill(start_color="DDECFF", end_color="DDECFF", fill_type="solid"),
    "2/8(일)": PatternFill(start_color="DDFFE0", end_color="DDFFE0", fill_type="solid"),
}


def create_daily_plan_sheet(ws, grouped_items):
    """Create Sheet 4: 일일계획 - 담당자별/날짜별 작업 배분표"""
    ws.title = "일일계획"

    # Collect all incomplete items with owner info
    all_items = []
    for group_name, items in grouped_items.items():
        for item in items:
            all_items.append({**item, "group": group_name})

    # Group by owner
    by_owner = OrderedDict()
    for item in all_items:
        owner = item["owner"] or "미배정"
        if owner not in by_owner:
            by_owner[owner] = []
        by_owner[owner].append(item)

    # Sort owners by item count (descending)
    by_owner = OrderedDict(sorted(by_owner.items(), key=lambda x: -len(x[1])))

    last_col_letter = "K"

    # Title
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'].value = "S25016 잔여 펀치리스트 일일 진행계획 (2/5~2/8)"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells(f'A2:{last_col_letter}2')
    today = datetime.now().strftime('%Y년 %m월 %d일')
    ws['A2'].value = f"{today} 작성 | 미완료 {len(all_items)}건 | 담당자 {len(by_owner)}명"
    ws['A2'].font = NORMAL_FONT
    ws['A2'].alignment = CENTER

    # Summary table: 담당자별 건수 x 우선순위
    row = 4
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].value = "■ 담당자별 미완료 현황"
    ws[f'A{row}'].font = SUBTITLE_FONT

    row += 1
    sum_headers = ["담당자", "전체", "긴급", "높음", "보통", "낮음", "주요 기능영역"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)

    for owner, items in by_owner.items():
        row += 1
        pcnt = {"긴급": 0, "높음": 0, "보통": 0, "낮음": 0}
        groups_set = set()
        for i in items:
            p = i["priority"] or "보통"
            pcnt[p] = pcnt.get(p, 0) + 1
            groups_set.add(i["group"])
        groups_str = ", ".join(sorted(groups_set)[:3])

        for col, val in enumerate([owner, len(items), pcnt["긴급"], pcnt["높음"],
                                    pcnt["보통"], pcnt["낮음"], groups_str], 1):
            cell = ws.cell(row=row, column=col, value=val)
            align = LEFT if col == 7 else CENTER
            style(cell, NORMAL_FONT, None, align, THIN_BORDER)

    # Column widths for plan section
    col_widths = {'A': 5, 'B': 14, 'C': 16, 'D': 42, 'E': 8, 'F': 8, 'G': 10,
                  'H': 16, 'I': 16, 'J': 16, 'K': 16}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    # Daily plan section: per owner
    row += 2
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'].value = "■ 담당자별 일일 작업 배분"
    ws[f'A{row}'].font = SUBTITLE_FONT

    row += 1
    plan_headers = ["No.", "기능영역", "ID", "항목명", "라인", "우선순위", "현재상태",
                    "2/5(목)", "2/6(금)", "2/7(토)", "2/8(일)"]
    for col, h in enumerate(plan_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        style(cell, HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER)
    header_row = row

    row += 1
    seq = 0

    for owner, items in by_owner.items():
        # Owner separator row
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        cell = ws.cell(row=row, column=1,
                       value=f"▶ {owner} ({len(items)}건)")
        ofill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        style(cell, Font(name="맑은 고딕", size=11, bold=True), ofill, LEFT, THIN_BORDER)
        ws.row_dimensions[row].height = 25
        row += 1

        # Sort: priority first
        sorted_items = sorted(items, key=lambda x: (
            PRIORITY_ORDER.get(x["priority"], 99),
            x["sort_key"][0], x["sort_key"][1]
        ))

        for item in sorted_items:
            seq += 1

            # Plan columns based on group membership
            group = item["group"]
            if group in GROUP_SCHEDULE_DAILY:
                sched = GROUP_SCHEDULE_DAILY[group]
                plan_vals = [sched.get(d, "") for d in PLAN_DAYS]
            else:
                plan_vals = ["이벤트성"] * 4

            data = [
                seq, item["group"], item["id"], item["title"],
                item["line"], item["priority"], item["status"],
                *plan_vals,
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                align = CENTER if col in (1, 5, 6, 7) else LEFT
                font = NORMAL_FONT
                fill = None

                if col == 6:  # priority
                    p = item["priority"]
                    if p in PRIORITY_FILLS:
                        fill = PRIORITY_FILLS[p]
                    if p in PRIORITY_FONTS:
                        font = PRIORITY_FONTS[p]

                if col == 7:  # status
                    fill = PatternFill(
                        start_color=STATUS_COLORS.get(item["status"], "FFFFFF"),
                        end_color=STATUS_COLORS.get(item["status"], "FFFFFF"),
                        fill_type="solid")
                    font = BOLD_FONT

                # Plan day columns
                if col == 8:
                    fill = PLAN_DAY_FILLS["2/5(목)"]
                elif col == 9:
                    fill = PLAN_DAY_FILLS["2/6(금)"]
                elif col == 10:
                    fill = PLAN_DAY_FILLS["2/7(토)"]
                elif col == 11:
                    fill = PLAN_DAY_FILLS["2/8(일)"]

                style(cell, font, fill, align, THIN_BORDER)

            ws.row_dimensions[row].height = 26
            row += 1

    # Data Validation for plan columns
    dv = DataValidation(
        type="list",
        formula1='"작업예정,디버깅,테스트,배포,확인,분석,이벤트성,미진행,-"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "목록에서 선택하세요"
    ws.add_data_validation(dv)
    dv.add(f"H{header_row + 1}:K{row - 1}")

    # Footer
    row += 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'].value = "2/5~2/8 열에 일일 작업계획을 입력하세요 (드롭다운 선택 가능)"
    ws[f'A{row}'].font = Font(name="맑은 고딕", size=10, italic=True, color="666666")
    ws[f'A{row}'].alignment = LEFT

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1


# === Main ===

def main():
    print("=== S25016 펀치리스트 엑셀 생성 ===\n")

    items = build_items()
    sorted_items = sort_items(items)
    total = len(sorted_items)

    # Filter and group incomplete items
    incomplete = filter_incomplete(items)
    grouped = group_and_sort_incomplete(incomplete)

    # Status summary
    scnt = {}
    for i in sorted_items:
        scnt[i["status"]] = scnt.get(i["status"], 0) + 1

    print(f"\n총 {total}건")
    print("\n상태별 분류:")
    for s in ["신규", "진행중", "검증중", "완료", "보류"]:
        print(f"  {s}: {scnt.get(s, 0)}건")
    print(f"  ─────────")
    print(f"  합계: {sum(scnt.values())}건")

    print(f"\n미완료 항목: {len(incomplete)}건")
    print("기능영역별:")
    for g, items_g in grouped.items():
        print(f"  {g}: {len(items_g)}건")

    # Create workbook
    wb = Workbook()

    # Sheet 1: 미완료 항목 + 계획
    ws_plan = wb.active
    create_plan_sheet(ws_plan, grouped)

    # Sheet 2: 요약
    ws_summary = wb.create_sheet()
    create_summary_sheet(ws_summary, sorted_items)

    # Sheet 3: 전체 목록
    ws_list = wb.create_sheet()
    create_list_sheet(ws_list, sorted_items)

    # Sheet 4: 일일계획
    ws_daily = wb.create_sheet()
    create_daily_plan_sheet(ws_daily, grouped)

    wb.save(OUTPUT_PATH)
    print(f"\n엑셀 파일 생성: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
